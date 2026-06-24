"""
Steganalysis Ringkas: NC, Entropy, BER, dan BER After Attack.
Fokus pada metrik utama + robustness testing + export Excel.

Metrik:
  1. Entropy           - Perbandingan entropi cover vs stego (semakin KECIL selisih semakin bagus)
  2. NC                - Normalized Correlation cover vs stego (semakin BESAR semakin bagus, ideal = 1.0)
  3. BER               - Bit Error Rate payload asli vs hasil ekstraksi (semakin KECIL semakin bagus, ideal = 0%)
  4. BER After Attack  - BER setelah serangan (Gaussian noise, low-pass, resample, amplitude scaling)

Usage:
    python steganalysis2.py
"""

import numpy as np
import scipy.io.wavfile as scp
from scipy.signal import butter, lfilter
from collections import Counter
import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import math
import io
from contextlib import redirect_stdout


# ==============================================================================
# UTILITY
# ==============================================================================

def load_audio_samples(filepath):
    rate, data = scp.read(filepath)
    if data.ndim > 1:
        data = data[:, 0]
    return rate, data.astype(np.float64)


def _build_interpolated_reference(cover_samples):
    """
    Rekonstruksi sinyal referensi yang sejajar dengan stego audio.
    Stego audio menyisipkan sampel interpolasi di antara sampel asli,
    sehingga panjangnya ~2N-1. Fungsi ini membuat versi 'bersih' (tanpa
    embedding) agar perbandingan metrik bisa dilakukan secara adil.
    """
    n = len(cover_samples)
    ref = np.zeros(2 * n - 1)
    ref[::2] = cover_samples
    ref[1::2] = np.floor((cover_samples[:-1] + cover_samples[1:]) / 2.0)
    return ref


def read_payload(file_payload):
    binary_data = list(open(file_payload))[0]
    binary_data = binary_data.split('\t')
    binary_data = [x.strip('ÿþ') for x in binary_data]
    binary_data = [x.strip('\x00') for x in binary_data]
    binary_data = ''.join(binary_data)
    return binary_data


# ==============================================================================
# 1. ENTROPY
# ==============================================================================

def calculate_entropy(samples):
    """
    Shannon Entropy: H(X) = -sum(p(x) * log2(p(x)))
    Semakin KECIL selisih entropy cover vs stego, semakin BAGUS.
    Nilai ideal: selisih = 0 (distribusi identik).
    """
    samples_int = samples.astype(np.int64)
    counts = Counter(samples_int)
    total = len(samples_int)
    entropy = 0.0
    for count in counts.values():
        p = count / total
        if p > 0:
            entropy -= p * math.log2(p)
    return entropy


def entropy_analysis(cover_samples, stego_samples):
    if len(stego_samples) >= len(cover_samples) * 1.8:
        ref = _build_interpolated_reference(cover_samples)
    else:
        ref = cover_samples
    e_cover = calculate_entropy(ref)
    e_stego = calculate_entropy(stego_samples)
    delta = abs(e_stego - e_cover)
    relative_change = (delta / e_cover * 100) if e_cover > 0 else 0
    return {
        'entropy_cover': e_cover,
        'entropy_stego': e_stego,
        'entropy_diff': delta,
        'entropy_pct': relative_change
    }


# ==============================================================================
# 2. NORMALIZED CORRELATION (NC)
# ==============================================================================

def normalized_correlation(cover_samples, stego_samples):
    """
    NC = sum(C*S) / sqrt(sum(C^2) * sum(S^2))
    Semakin BESAR NC semakin BAGUS. Nilai ideal: NC = 1.0.

    Menangani kasus stego hasil interpolasi (panjang ~2x cover) dengan
    membangun sinyal referensi yang sejajar.
    """
    if len(stego_samples) >= len(cover_samples) * 1.8:
        c = _build_interpolated_reference(cover_samples)
    else:
        c = cover_samples

    min_len = min(len(c), len(stego_samples))
    c = c[:min_len]
    s = stego_samples[:min_len]
    numerator = np.sum(c * s)
    denominator = np.sqrt(np.sum(c ** 2) * np.sum(s ** 2))
    nc = numerator / denominator if denominator > 0 else 0
    return nc


def evaluate_audio_pair(cover_path, stego_path):
    """Evaluate Entropy and NC for one selected cover/stego pair.

    This deliberately excludes BER because BER needs the original and extracted
    payloads, while the GUI evaluation form only receives two audio files.
    """
    if not os.path.isfile(cover_path):
        raise FileNotFoundError(f"Cover audio tidak ditemukan: {cover_path}")
    if not os.path.isfile(stego_path):
        raise FileNotFoundError(f"Stego audio tidak ditemukan: {stego_path}")

    _, cover = load_audio_samples(cover_path)
    _, stego = load_audio_samples(stego_path)

    entropy = entropy_analysis(cover, stego)
    return {
        **entropy,
        'nc': normalized_correlation(cover, stego),
    }


# ==============================================================================
# 3. BER (Bit Error Rate)
# ==============================================================================

def calculate_ber(original_binary, extracted_binary):
    """
    BER = (jumlah bit salah / total bit) x 100%
    Semakin KECIL BER semakin BAGUS. Nilai ideal: BER = 0%.
    """
    def to_binary_str(data):
        if isinstance(data, (list, tuple)):
            return ''.join(str(b) for b in data if str(b) in '01')
        s = str(data).replace(' ', '').replace('\t', '')
        return ''.join(c for c in s if c in '01')

    original_binary = to_binary_str(original_binary)
    extracted_binary = to_binary_str(extracted_binary)
    min_length = min(len(original_binary), len(extracted_binary))

    if min_length == 0:
        return 100.0, 0, 0

    bit_errors = sum(1 for i in range(min_length) if original_binary[i] != extracted_binary[i])
    ber_percentage = (bit_errors / min_length) * 100
    return ber_percentage, bit_errors, min_length


# ==============================================================================
# 4. ATTACKS & BER AFTER ATTACK (Robustness)
# ==============================================================================

DEFAULT_ATTACKS = {
    'gaussian_noise_30dB': {'type': 'noise', 'snr_db': 30},
    'gaussian_noise_20dB': {'type': 'noise', 'snr_db': 20},
    'lowpass_4000Hz': {'type': 'lowpass', 'cutoff': 4000},
    'lowpass_2000Hz': {'type': 'lowpass', 'cutoff': 2000},
    'resampling_22050': {'type': 'resample', 'target_rate': 22050},
    'amplitude_90pct': {'type': 'amplitude', 'factor': 0.9},
    'amplitude_80pct': {'type': 'amplitude', 'factor': 0.8},
}

ATTACK_LABELS = {
    'gaussian_noise_30dB': 'Noise 30dB',
    'gaussian_noise_20dB': 'Noise 20dB',
    'lowpass_4000Hz': 'LPF 4kHz',
    'lowpass_2000Hz': 'LPF 2kHz',
    'resampling_22050': 'Resample 22k',
    'amplitude_90pct': 'Amp 90%',
    'amplitude_80pct': 'Amp 80%',
}


def _apply_attack(samples, rate, params):
    t = params['type']
    if t == 'noise':
        power = np.mean(samples ** 2)
        noise_power = power / (10 ** (params['snr_db'] / 10))
        return samples + np.random.normal(0, np.sqrt(noise_power), len(samples))
    elif t == 'lowpass':
        nyq = 0.5 * rate
        norm_cut = min(params['cutoff'] / nyq, 0.99)
        b, a = butter(5, norm_cut, btype='low')
        return lfilter(b, a, samples)
    elif t == 'resample':
        dur = len(samples) / rate
        n_t = int(dur * params['target_rate'])
        x_orig = np.linspace(0, dur, len(samples))
        x_tgt = np.linspace(0, dur, n_t)
        down = np.interp(x_tgt, x_orig, samples)
        return np.interp(np.linspace(0, dur, len(samples)), x_tgt, down)
    elif t == 'amplitude':
        return samples * params['factor']
    return samples


def _extract_payload(stego_base, share_indices):
    """Extract payload dari stego audio via Shamir reconstruction."""
    from methods import (extraction_sampling, separate,
                         extraction_interpolation_linear,
                         extraction_determine_selisih,
                         extraction_determine_sample_space,
                         reconstruct_secret, decimal_to_binary)
    devnull = io.StringIO()
    try:
        with redirect_stdout(devnull):
            _, stego_sample = extraction_sampling(stego_base, share_indices)
            orig, emb = separate(stego_sample)
            interp = extraction_interpolation_linear(orig)
            result = extraction_determine_selisih(emb, interp)
            if result is None:
                return None
            data_sel, last_idx, prime, share_no, last_bit = result
            bit = extraction_determine_sample_space(interp, last_idx)
            dec_payload = reconstruct_secret(data_sel, prime, share_no)
            return decimal_to_binary(dec_payload, bit, last_bit)
    except Exception:
        return None


def run_ber_after_attack(audio_no, payload_no, n_shares, min_shares, attacks=None):
    """
    Apply attacks → extract → compute BER untuk satu kombinasi audio-payload.
    Returns: dict {attack_name: ber_pct} atau None jika gagal.
    """
    if attacks is None:
        attacks = DEFAULT_ATTACKS

    orig_path = f'stegoaudioDataset/Payload/payload{payload_no}.txt'
    if not os.path.exists(orig_path):
        return None

    orig_bin = read_payload(orig_path)
    stego_base = f'stego_audio/stego_audio{audio_no}_payload{payload_no}/stegoaudio'
    attacked_base = f'attacked_audio/stego_audio{audio_no}_payload{payload_no}'
    share_indices = list(range(min_shares))

    for s in share_indices:
        if not os.path.exists(stego_base + str(s) + '.wav'):
            return None

    results = {}
    for atk_name, params in attacks.items():
        out_dir = os.path.join(attacked_base, atk_name)
        os.makedirs(out_dir, exist_ok=True)

        for s in share_indices:
            rate, samples = load_audio_samples(stego_base + str(s) + '.wav')
            attacked = _apply_attack(samples, rate, params)
            out_path = os.path.join(out_dir, f'stegoaudio{s}.wav')
            scp.write(out_path, rate,
                      np.clip(attacked, -32768, 32767).astype(np.int16))

        atk_stego_base = os.path.join(attacked_base, atk_name, 'stegoaudio')
        extracted = _extract_payload(atk_stego_base, share_indices)

        if extracted is not None:
            ber_pct, _, _ = calculate_ber(orig_bin, extracted)
        else:
            ber_pct = None
        results[atk_name] = ber_pct

    return results


# ==============================================================================
# DETECT SHARES
# ==============================================================================

def _detect_total_shares(audio_no, payload_no):
    idx = 0
    while os.path.exists(f'stego_audio/stego_audio{audio_no}_payload{payload_no}/stegoaudio{idx}.wav'):
        idx += 1
    return idx


# ==============================================================================
# EXCEL STYLING
# ==============================================================================

def _style_header(ws, row, max_col):
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border


def _dc(cell, fmt=None):
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt


PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ==============================================================================
# EXPORT EXCEL
# ==============================================================================

def export_to_excel(all_rows, all_aggregate, excel_path, attack_ber_data=None):
    """
    Export ke Excel dengan sheet:
      1. Detail Per-Share          : Audio, Payload, Share, Entropy, NC, BER
      2. Ringkasan (Avg)           : Rata-rata per Audio-Payload
      3. Ringkasan (Best)          : Nilai terbaik per Audio-Payload
      4. Ringkasan (Worst)         : Nilai terburuk per Audio-Payload
      5+. Pivot per metrik per share : Matriks Audio x Payload (siap grafik)
      +  BER After Attack          : Pivot BER per serangan
      Terakhir. Keterangan Metrik  : Teori & threshold
    """
    wb = xl.Workbook()
    wb.remove(wb.active)

    # ----- Sheet 1: Detail Per-Share -----
    ws1 = wb.create_sheet("Detail Per-Share")
    headers = [
        "Audio", "Payload", "Share",
        "Entropy Cover (bits)", "Entropy Stego (bits)",
        "Entropy Diff (bits)", "Entropy Change (%)",
        "NC",
        "BER (%)", "Bit Errors", "Total Bits"
    ]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    _style_header(ws1, 1, len(headers))

    for r, rd in enumerate(all_rows, 2):
        m = rd['m']
        vals = [
            rd['audio'], rd['payload'], rd['share'],
            m['entropy_cover'], m['entropy_stego'],
            m['entropy_diff'], m['entropy_pct'],
            m['nc'],
            m['ber_pct'], m['ber_errors'], m['ber_total']
        ]
        fmts = [
            None, None, None,
            '0.000000', '0.000000',
            '0.000000', '0.0000',
            '0.0000000000',
            '0.0000', '0', '0'
        ]
        for c, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws1.cell(row=r, column=c, value=v)
            _dc(cell, f)

    for c in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 20

    # ----- Sheets 2-4: Ringkasan Avg / Best / Worst -----
    summary_cols = [
        ("Entropy Change (%)", 'entropy_pct', '0.0000', 'min'),
        ("NC", 'nc', '0.0000000000', 'max'),
        ("BER (%)", 'ber_pct', '0.0000', 'min'),
    ]

    for label, agg_type in [("Ringkasan (Avg)", 'avg'), ("Ringkasan (Best)", 'best'), ("Ringkasan (Worst)", 'worst')]:
        ws = wb.create_sheet(label)
        sh = ["Audio", "Payload", "#Share"] + [s[0] for s in summary_cols]
        for c, h in enumerate(sh, 1):
            ws.cell(row=1, column=c, value=h)
        _style_header(ws, 1, len(sh))

        for r, agg in enumerate(all_aggregate, 2):
            ws.cell(row=r, column=1, value=agg['audio']); _dc(ws.cell(row=r, column=1))
            ws.cell(row=r, column=2, value=agg['payload']); _dc(ws.cell(row=r, column=2))
            ws.cell(row=r, column=3, value=agg['n_shares']); _dc(ws.cell(row=r, column=3))

            for ci, (_, key, fmt, direction) in enumerate(summary_cols):
                vals = [m[key] for m in agg['metrics']]
                if agg_type == 'avg':
                    val = np.mean(vals)
                elif agg_type == 'best':
                    val = max(vals) if direction == 'max' else min(vals)
                else:
                    val = min(vals) if direction == 'max' else max(vals)
                cell = ws.cell(row=r, column=4 + ci, value=val)
                _dc(cell, fmt)

        for c in range(1, len(sh) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 20

    # ----- Pivot sheets per metrik per share -----
    all_shares = sorted(set(r['share'] for r in all_rows))
    all_audios = sorted(set(r['audio'] for r in all_rows))
    all_payloads = sorted(set(r['payload'] for r in all_rows))
    lookup = {(r['audio'], r['payload'], r['share']): r['m'] for r in all_rows}

    pivot_defs = [
        ("Entropy (%)", 'entropy_pct', '0.0000', 1, 'min'),
        ("NC", 'nc', '0.0000000000', 0.999, 'max'),
        ("BER (%)", 'ber_pct', '0.0000', 1, 'min'),
    ]

    for metric_label, key, fmt, threshold, direction in pivot_defs:
        for s in all_shares:
            sname = f"{metric_label} S{s}"
            if len(sname) > 31:
                sname = sname[:31]
            ws = wb.create_sheet(sname)

            ws.cell(row=1, column=1, value=f"{metric_label} (Share {s})")
            ws.cell(row=1, column=1).font = Font(bold=True, size=12)

            ws.cell(row=2, column=1, value="Audio \\ Payload")
            for ci, p in enumerate(all_payloads):
                ws.cell(row=2, column=ci + 2, value=f"Payload {p}")
            _style_header(ws, 2, len(all_payloads) + 1)

            for ri, a in enumerate(all_audios):
                lbl = ws.cell(row=ri + 3, column=1, value=f"Audio {a}")
                lbl.font = Font(bold=True)
                _dc(lbl)

                for ci, p in enumerate(all_payloads):
                    m = lookup.get((a, p, s))
                    if m:
                        val = m[key]
                        cell = ws.cell(row=ri + 3, column=ci + 2, value=val)
                        _dc(cell, fmt)
                        if direction == 'min':
                            cell.fill = PASS_FILL if val < threshold else FAIL_FILL
                        else:
                            cell.fill = PASS_FILL if val >= threshold else FAIL_FILL
                    else:
                        cell = ws.cell(row=ri + 3, column=ci + 2, value="-")
                        _dc(cell)

            ws.column_dimensions['A'].width = 14
            for ci in range(len(all_payloads)):
                ws.column_dimensions[get_column_letter(ci + 2)].width = 14

    # ----- Sheet BER After Attack -----
    if attack_ber_data:
        atk_names = list(DEFAULT_ATTACKS.keys())
        ws_atk = wb.create_sheet("BER After Attack")
        atk_headers = ["Audio", "Payload"] + [ATTACK_LABELS.get(a, a) for a in atk_names]
        for c, h in enumerate(atk_headers, 1):
            ws_atk.cell(row=1, column=c, value=h)
        _style_header(ws_atk, 1, len(atk_headers))

        for r, ad in enumerate(attack_ber_data, 2):
            ws_atk.cell(row=r, column=1, value=ad['audio']); _dc(ws_atk.cell(row=r, column=1))
            ws_atk.cell(row=r, column=2, value=ad['payload']); _dc(ws_atk.cell(row=r, column=2))
            for ci, atk in enumerate(atk_names):
                val = ad['ber'].get(atk)
                if val is None:
                    cell = ws_atk.cell(row=r, column=3 + ci, value="GAGAL")
                    _dc(cell)
                    cell.fill = FAIL_FILL
                else:
                    cell = ws_atk.cell(row=r, column=3 + ci, value=val)
                    _dc(cell, '0.0000')
                    if val < 1:
                        cell.fill = PASS_FILL
                    elif val < 5:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = FAIL_FILL

        for c in range(1, len(atk_headers) + 1):
            ws_atk.column_dimensions[get_column_letter(c)].width = 16

    # ----- Sheet Keterangan -----
    ws_k = wb.create_sheet("Keterangan Metrik")
    info = [
        ["Metrik", "Arah Nilai Baik", "Threshold PASS", "Nilai Ideal", "Rumus / Keterangan"],
        ["Entropy Change (%)",
         "Semakin KECIL semakin bagus", "< 1%", "0%",
         "Selisih relatif Shannon Entropy cover vs stego. H(X) = -sum(p(x)*log2(p(x)))"],
        ["NC (Normalized Correlation)",
         "Semakin BESAR semakin bagus", ">= 0.999", "1.0",
         "Korelasi sinyal. NC = sum(C*S) / sqrt(sum(C^2)*sum(S^2)). Rentang: -1 s/d 1"],
        ["BER (Bit Error Rate)",
         "Semakin KECIL semakin bagus", "< 1%", "0%",
         "Perbandingan bit payload asli vs hasil ekstraksi. BER = (bit salah / total bit) x 100%"],
        ["BER After Attack",
         "Semakin KECIL semakin bagus", "< 5%", "0%",
         "BER setelah serangan (noise, filter, resample, amplitude). Mengukur robustness payload."],
        ["", "", "", "",
         "Serangan: Gaussian Noise 30/20dB, Low-Pass 4k/2kHz, Resample 22050Hz, Amplitude 90/80%"],
    ]
    for ri, row in enumerate(info):
        for ci, val in enumerate(row):
            cell = ws_k.cell(row=ri + 1, column=ci + 1, value=val)
            _dc(cell)
    _style_header(ws_k, 1, 5)
    for c in range(1, 6):
        ws_k.column_dimensions[get_column_letter(c)].width = 32

    os.makedirs(os.path.dirname(excel_path) if os.path.dirname(excel_path) else '.', exist_ok=True)
    wb.save(excel_path)
    print(f"\n  [OK] Hasil disimpan ke: {excel_path}")


# ==============================================================================
# SINGLE ANALYSIS
# ==============================================================================

def run_single(audio_no='1', payload_no='1', stego_share=0,
               min_shares=None, total_shares_override=None):
    cover_path = f'stegoaudioDataset/Audio/data{audio_no}_mono.wav'
    stego_path = f'stego_audio/stego_audio{audio_no}_payload{payload_no}/stegoaudio{stego_share}.wav'
    original_payload_path = f'stegoaudioDataset/Payload/payload{payload_no}.txt'
    extracted_payload_path = f'extracted/stego_audio{audio_no}_payload{payload_no}/payload.txt'

    print("=" * 70)
    print("   STEGANALYSIS (NC, Entropy, BER, BER After Attack)")
    print(f"   Cover : {cover_path}")
    print(f"   Stego : {stego_path}")
    print("=" * 70)

    if not os.path.exists(cover_path):
        print(f"[ERROR] Cover audio tidak ditemukan: {cover_path}"); return
    if not os.path.exists(stego_path):
        print(f"[ERROR] Stego audio tidak ditemukan: {stego_path}"); return

    _, cover = load_audio_samples(cover_path)
    _, stego = load_audio_samples(stego_path)

    # --- Entropy ---
    print("\n" + "-" * 70)
    print("  1. ENTROPY ANALYSIS")
    print("-" * 70)
    print("  Teori: Semakin KECIL selisih entropy, semakin BAGUS (ideal = 0).")
    print("    < 1% = aman | 1-5% = cukup | > 5% = terdeteksi")
    ent = entropy_analysis(cover, stego)
    print(f"  Entropy cover  : {ent['entropy_cover']:.6f} bits")
    print(f"  Entropy stego  : {ent['entropy_stego']:.6f} bits")
    print(f"  Selisih        : {ent['entropy_diff']:.6f} bits")
    print(f"  Perubahan      : {ent['entropy_pct']:.4f}%")

    # --- NC ---
    print("\n" + "-" * 70)
    print("  2. NORMALIZED CORRELATION (NC)")
    print("-" * 70)
    print("  Teori: Semakin BESAR NC, semakin BAGUS (ideal = 1.0).")
    print("    >= 0.9999 = identik | >= 0.999 = sangat baik | >= 0.99 = baik")
    nc = normalized_correlation(cover, stego)
    print(f"  NC             : {nc:.10f}")

    # --- BER ---
    print("\n" + "-" * 70)
    print("  3. BER (Bit Error Rate)")
    print("-" * 70)
    print("  Teori: Semakin KECIL BER, semakin BAGUS (ideal = 0%).")
    print("    0% = sempurna | < 1% = sangat baik | < 5% = baik")

    if os.path.exists(original_payload_path) and os.path.exists(extracted_payload_path):
        orig = read_payload(original_payload_path)
        extr = read_payload(extracted_payload_path)
        ber_pct, ber_err, ber_total = calculate_ber(orig, extr)
        print(f"  Total bit      : {ber_total}")
        print(f"  Bit error      : {ber_err}")
        print(f"  BER            : {ber_pct:.4f}%")
    else:
        print("  [SKIP] File payload asli atau hasil ekstraksi tidak ditemukan.")
        print(f"    Original  : {original_payload_path}")
        print(f"    Extracted : {extracted_payload_path}")

    # --- BER After Attack ---
    print("\n" + "-" * 70)
    print("  4. BER AFTER ATTACK (Robustness)")
    print("-" * 70)
    print("  Teori: Semakin KECIL BER setelah serangan, semakin ROBUST.")
    print("    0% = sempurna | < 1% = sangat baik | < 5% = baik | >= 5% = perlu perhatian")

    if min_shares is not None:
        n_sh = total_shares_override if total_shares_override else _detect_total_shares(
            int(audio_no), int(payload_no))
        if n_sh >= min_shares:
            print(f"  Menjalankan 7 serangan (n={n_sh}, k={min_shares}) ...")
            atk_result = run_ber_after_attack(
                int(audio_no), int(payload_no), n_sh, min_shares)
            if atk_result:
                for atk_name, ber_val in atk_result.items():
                    label = ATTACK_LABELS.get(atk_name, atk_name)
                    if ber_val is not None:
                        print(f"  {label:<20}: {ber_val:.4f}%")
                    else:
                        print(f"  {label:<20}: GAGAL (ekstraksi gagal)")
            else:
                print("  [SKIP] Tidak bisa menjalankan serangan.")
        else:
            print(f"  [SKIP] Jumlah share ({n_sh}) < min_shares ({min_shares}).")
    else:
        print("  [SKIP] min_shares tidak ditentukan. Tambahkan parameter k untuk tes serangan.")

    print("=" * 70)


# ==============================================================================
# BATCH ANALYSIS + EXCEL
# ==============================================================================

def run_batch(total_audio=15, total_payload=11, total_shares=None,
              min_shares=None, excel_output=None, run_attacks=True):
    all_rows = []
    all_aggregate = []

    print("=" * 100)
    print("   BATCH STEGANALYSIS (NC, Entropy, BER) - semua share")
    print("=" * 100)

    header = (f"{'Audio':>6} | {'Payload':>7} | {'Share':>5} | "
              f"{'Entropy%':>9} | {'NC':>14} | {'BER%':>8}")
    print(header)
    print("-" * 60)

    for x in range(1, total_audio + 1):
        for y in range(1, total_payload + 1):
            cover_path = f'stegoaudioDataset/Audio/data{x}_mono.wav'
            if not os.path.exists(cover_path):
                continue

            n_shares = total_shares if total_shares else _detect_total_shares(x, y)
            if n_shares == 0:
                continue

            _, cover = load_audio_samples(cover_path)

            original_payload_path = f'stegoaudioDataset/Payload/payload{y}.txt'
            extracted_payload_path = f'extracted/stego_audio{x}_payload{y}/payload.txt'

            has_ber = os.path.exists(original_payload_path) and os.path.exists(extracted_payload_path)
            if has_ber:
                orig_bin = read_payload(original_payload_path)
                extr_bin = read_payload(extracted_payload_path)
                ber_pct, ber_err, ber_total = calculate_ber(orig_bin, extr_bin)
            else:
                ber_pct, ber_err, ber_total = None, None, None

            combo_metrics = []

            for s in range(n_shares):
                stego_path = f'stego_audio/stego_audio{x}_payload{y}/stegoaudio{s}.wav'
                if not os.path.exists(stego_path):
                    continue

                _, stego = load_audio_samples(stego_path)
                ent = entropy_analysis(cover, stego)
                nc = normalized_correlation(cover, stego)

                m = {
                    'entropy_cover': ent['entropy_cover'],
                    'entropy_stego': ent['entropy_stego'],
                    'entropy_diff': ent['entropy_diff'],
                    'entropy_pct': ent['entropy_pct'],
                    'nc': nc,
                    'ber_pct': ber_pct if ber_pct is not None else 0,
                    'ber_errors': ber_err if ber_err is not None else 0,
                    'ber_total': ber_total if ber_total is not None else 0,
                }
                combo_metrics.append(m)
                all_rows.append({'audio': x, 'payload': y, 'share': s, 'm': m})

                ber_str = f"{ber_pct:>7.4f}%" if ber_pct is not None else "    N/A"
                row = (f"{x:>6} | {y:>7} | {s:>5} | "
                       f"{m['entropy_pct']:>8.4f}% | "
                       f"{nc:>14.10f} | "
                       f"{ber_str}")
                print(row)

            if combo_metrics:
                all_aggregate.append({
                    'audio': x, 'payload': y,
                    'n_shares': len(combo_metrics),
                    'metrics': combo_metrics
                })

    print("=" * 100)

    if not all_aggregate:
        print("\n[INFO] Tidak ada data stego audio yang ditemukan.")
        return

    # --- Ringkasan ---
    print("\n" + "=" * 100)
    print("   RINGKASAN PER KOMBINASI (avg / best / worst)")
    print("   Entropy%: best=min, worst=max | NC: best=max, worst=min | BER%: best=min, worst=max")
    print("=" * 100)

    header2 = (f"{'Audio':>6} | {'Payload':>7} | {'#Sh':>4} | "
               f"{'Ent% avg':>9} | {'Ent% best':>9} | {'Ent% worst':>10} | "
               f"{'NC avg':>14} | {'NC best':>14} | {'NC worst':>14} | "
               f"{'BER%':>7}")
    print(header2)
    print("-" * 100)

    for agg in all_aggregate:
        ms = agg['metrics']
        ent_vals = [m['entropy_pct'] for m in ms]
        nc_vals = [m['nc'] for m in ms]
        ber_val = ms[0]['ber_pct']

        row2 = (f"{agg['audio']:>6} | {agg['payload']:>7} | {agg['n_shares']:>4} | "
                f"{np.mean(ent_vals):>8.4f}% | {min(ent_vals):>8.4f}% | {max(ent_vals):>9.4f}% | "
                f"{np.mean(nc_vals):>14.10f} | {max(nc_vals):>14.10f} | {min(nc_vals):>14.10f} | "
                f"{ber_val:>6.4f}%")
        print(row2)

    print("=" * 100)

    # --- BER After Attack ---
    attack_ber_data = []
    if run_attacks and min_shares is not None:
        print("\n" + "=" * 120)
        print("   BER AFTER ATTACK (Robustness Test)")
        print("   Menerapkan serangan → ekstraksi ulang → hitung BER")
        print("=" * 120)

        atk_names = list(DEFAULT_ATTACKS.keys())
        atk_header = f"{'Audio':>6} | {'Payload':>7} | " + " | ".join(
            f"{ATTACK_LABELS.get(a, a):>12}" for a in atk_names)
        print(atk_header)
        print("-" * 120)

        total_combos = len(all_aggregate)
        for idx, agg in enumerate(all_aggregate, 1):
            x, y = agg['audio'], agg['payload']
            n_sh = agg['n_shares']

            print(f"  [{idx}/{total_combos}] Audio {x}, Payload {y} ... ", end="", flush=True)
            atk_result = run_ber_after_attack(x, y, n_sh, min_shares)

            if atk_result:
                attack_ber_data.append({'audio': x, 'payload': y, 'ber': atk_result})
                vals_str = " | ".join(
                    f"{atk_result[a]:>11.4f}%" if atk_result[a] is not None else "       GAGAL"
                    for a in atk_names)
                print(f"\r{x:>6} | {y:>7} | {vals_str}")
            else:
                print(f"\r{x:>6} | {y:>7} | {'SKIP':>12}" + " | " * (len(atk_names) - 1))

        print("=" * 120)

    # --- Export Excel ---
    if excel_output is None:
        n_str = str(total_shares) if total_shares else 'auto'
        excel_output = f'steganalysis_results/steganalysis2_{n_str}_shares.xlsx'

    export_to_excel(all_rows, all_aggregate, excel_output, attack_ber_data)
    return all_rows, all_aggregate


# ==============================================================================
# MAIN
# ==============================================================================

if __name__ == '__main__':
    import sys

    if len(sys.argv) >= 3:
        audio_no = sys.argv[1]
        payload_no = sys.argv[2]
        share_no = int(sys.argv[3]) if len(sys.argv) > 3 else 0
    else:
        audio_no = input("Nomor audio (default=1): ").strip() or '1'
        payload_no = input("Nomor payload (default=1): ").strip() or '1'
        share_input = input("Nomor share stego (default=0): ").strip()
        share_no = int(share_input) if share_input else 0

    mode = input(
        "\nPilih mode:\n"
        "  1. Single analysis (detail 1 pasangan)\n"
        "  2. Batch analysis (semua kombinasi) + Export Excel\n"
        "Pilihan (1/2, default=1): "
    ).strip() or '1'

    k_input = input("Minimum shares / k (untuk BER After Attack, kosong = skip): ").strip()
    k_shares = int(k_input) if k_input else None

    if mode == '2':
        shares_input = input("Total shares / n (kosong = auto-detect): ").strip()
        n_shares = int(shares_input) if shares_input else None
        run_batch(total_shares=n_shares, min_shares=k_shares)
    else:
        run_single(audio_no, payload_no, share_no, min_shares=k_shares)
