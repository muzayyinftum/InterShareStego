import os
import math
import numpy as np
import scipy.io.wavfile as scp
from scipy.fft import fft
from scipy.stats import skew, kurtosis
from sklearn.svm import SVC
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    roc_auc_score, confusion_matrix, classification_report
)
from sklearn.preprocessing import StandardScaler
from sklearn.utils import resample

# ==============================================================================
# CONFIGURATION
# ==============================================================================

COVER_PATH = 'DATASET/Audio/data1_mono.wav'
STEGO_BASE = 'STEGOAUDIO'
SEGMENT_LENGTH = 4096       # samples per segment (~93ms @ 44.1kHz)
SEGMENT_OVERLAP = 0.5      # 50% overlap
RANDOM_STATE = 42
TEST_SIZE = 0.3
BALANCE_CLASSES = True   # Undersample stego to balance classes for fairer evaluation


# ==============================================================================
# LOAD AUDIO
# ==============================================================================

def load_audio_samples(filepath):
    """Read WAV and return (rate, float64 samples)."""
    rate, data = scp.read(filepath)
    if data.ndim > 1:
        data = data[:, 0]
    return rate, data.astype(np.float64)


def build_interpolated_reference(cover_samples):
    n = len(cover_samples)
    ref = np.zeros(2 * n - 1)
    ref[::2] = cover_samples
    ref[1::2] = np.floor((cover_samples[:-1] + cover_samples[1:]) / 2.0)
    return ref


# ==============================================================================
# SEGMENTATION
# ==============================================================================

def get_segments(samples, seg_len, overlap_ratio=0.5):
    step = int(seg_len * (1 - overlap_ratio))
    if step < 1:
        step = 1
    segments = []
    for start in range(0, len(samples) - seg_len + 1, step):
        segments.append(samples[start:start + seg_len])
    return segments


# ==============================================================================
# FEATURE EXTRACTION
# ==============================================================================

def extract_statistical_features(segment):
    return [
        np.mean(segment),
        np.std(segment) if np.std(segment) > 0 else 1e-10,
        skew(segment),
        kurtosis(segment),
        np.median(segment),
        np.max(np.abs(segment)),
    ]


def extract_spectral_features(segment):
    fft_mag = np.abs(fft(segment))[:len(segment) // 2]
    if len(fft_mag) < 10:
        return [0.0] * 8

    n = len(fft_mag)
    band_size = n // 4
    bands = [
        fft_mag[:band_size],
        fft_mag[band_size:2*band_size],
        fft_mag[2*band_size:3*band_size],
        fft_mag[3*band_size:],
    ]
    features = []
    for band in bands:
        if len(band) > 0:
            features.append(np.mean(band))
            features.append(np.std(band) if np.std(band) > 0 else 1e-10)
        else:
            features.extend([0.0, 0.0])
    return features[:8]


def extract_difference_features(segment):
    diff = np.diff(segment.astype(np.int64))
    if len(diff) < 2:
        return [0.0] * 6
    return [
        np.mean(np.abs(diff)),
        np.std(diff) if np.std(diff) > 0 else 1e-10,
        skew(diff),
        np.median(np.abs(diff)),
        np.sum(diff == 0) / len(diff),  # proportion of diff=0
        np.sum(np.abs(diff) == 1) / len(diff),  # proportion of |diff|=1 (LSB-related)
    ]


def extract_histogram_moments(segment, n_bins=32):
    hist, _ = np.histogram(segment, bins=n_bins, density=True)
    hist = hist + 1e-10
    hist = hist / np.sum(hist)
    return [
        np.mean(hist),
        np.std(hist),
        skew(hist),
    ]


def extract_features_for_segment(segment):
    feats = []
    feats.extend(extract_statistical_features(segment))
    feats.extend(extract_spectral_features(segment))
    feats.extend(extract_difference_features(segment))
    feats.extend(extract_histogram_moments(segment))
    return np.array(feats, dtype=np.float64)


def extract_features_for_audio(samples, seg_len, overlap):
    segments = get_segments(samples, seg_len, overlap)
    return np.array([extract_features_for_segment(seg) for seg in segments])


# ==============================================================================
# BUILD DATASET
# ==============================================================================

def collect_cover_and_stego_files(cover_path=None):
    cover = cover_path or COVER_PATH
    stego_files = []

    # Search embedding output folders, e.g., stego_audio1_payload1.
    base_dir = STEGO_BASE
    if not os.path.exists(base_dir):
        return cover, stego_files

    for name in os.listdir(base_dir):
        if name.startswith('stego_audio'):
            folder = os.path.join(base_dir, name)
            if os.path.isdir(folder):
                for f in sorted(os.listdir(folder)):
                    if f.endswith('.wav'):
                        stego_files.append(os.path.join(folder, f))

    return cover, stego_files


def build_dataset(cover_path=None):
    cover_path, stego_paths = collect_cover_and_stego_files(cover_path)

    if not os.path.exists(cover_path):
        raise FileNotFoundError(f"audio cover not found: {cover_path}")

    if not stego_paths:
        raise FileNotFoundError(
            f"No stego files found. Please ensure folders like "
            f"STEGOAUDIO/stego_audio1_payload1/ contain stegoaudio*.wav"
        )

    rate_c, cover_raw = load_audio_samples(cover_path)

    cover_samples = build_interpolated_reference(cover_raw)

    X_list = []
    y_list = []

    # Cover segments -> label 0
    X_cover = extract_features_for_audio(cover_samples, SEGMENT_LENGTH, SEGMENT_OVERLAP)
    X_list.append(X_cover)
    y_list.append(np.zeros(len(X_cover), dtype=int))

    # Stego segments -> label 1
    for sp in stego_paths:
        rate_s, stego_samples = load_audio_samples(sp)

        min_len = min(len(cover_samples), len(stego_samples))
        if min_len < SEGMENT_LENGTH:
            continue
        stego_trunc = stego_samples[:min_len]
        X_stego = extract_features_for_audio(stego_trunc, SEGMENT_LENGTH, SEGMENT_OVERLAP)
        X_list.append(X_stego)
        y_list.append(np.ones(len(X_stego), dtype=int))

    X = np.vstack(X_list)
    y = np.concatenate(y_list)

    if BALANCE_CLASSES:
        idx_cover = np.where(y == 0)[0]
        idx_stego = np.where(y == 1)[0]
        n_cover = len(idx_cover)
        n_stego = len(idx_stego)
        if n_stego > n_cover:
            idx_stego_down = resample(
                idx_stego, replace=False, n_samples=n_cover, random_state=RANDOM_STATE
            )
            idx_keep = np.concatenate([idx_cover, idx_stego_down])
            np.random.RandomState(RANDOM_STATE).shuffle(idx_keep)
            X = X[idx_keep]
            y = y[idx_keep]

    return X, y, cover_path, stego_paths


def build_single_pair_dataset(cover_path, stego_path):
    """Build a balanced segment dataset from one GUI-selected audio pair."""
    if not os.path.isfile(cover_path):
        raise FileNotFoundError(f"Cover audio not found: {cover_path}")
    if not os.path.isfile(stego_path):
        raise FileNotFoundError(f"Stego audio not found: {stego_path}")

    _, cover_raw = load_audio_samples(cover_path)
    _, stego = load_audio_samples(stego_path)

    # The embedding method normally produces a 2N-1 stego signal.
    cover = (build_interpolated_reference(cover_raw)
             if len(stego) >= len(cover_raw) * 1.8 else cover_raw)
    min_len = min(len(cover), len(stego))
    if min_len < SEGMENT_LENGTH:
        raise ValueError(
            f"Audio too short for detector (minimum {SEGMENT_LENGTH} samples)."
        )

    X_cover = extract_features_for_audio(
        cover[:min_len], SEGMENT_LENGTH, SEGMENT_OVERLAP)
    X_stego = extract_features_for_audio(
        stego[:min_len], SEGMENT_LENGTH, SEGMENT_OVERLAP)
    pair_count = min(len(X_cover), len(X_stego))
    if pair_count < 3:
        raise ValueError("Not enough audio segments for training/testing the detector.")

    X = np.vstack((X_cover[:pair_count], X_stego[:pair_count]))
    X = np.nan_to_num(X, nan=0.0, posinf=0.0, neginf=0.0)
    y = np.concatenate((
        np.zeros(pair_count, dtype=int),
        np.ones(pair_count, dtype=int),
    ))
    return X, y


def build_cover_path(audio_no):
    return f'DATASET/Audio/data{audio_no}_mono.wav'


def evaluate_audio_pair(cover_path, stego_path):
    """Run the detector benchmark for one cover/stego pair and return metrics."""
    X, y = build_single_pair_dataset(cover_path, stego_path)
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y
    )

    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    clf = SVC(kernel='rbf', C=1.0, gamma='scale', random_state=RANDOM_STATE)
    clf.fit(X_train_scaled, y_train)
    y_pred = clf.predict(X_test_scaled)

    acc = accuracy_score(y_test, y_pred)
    prec = precision_score(y_test, y_pred, zero_division=0)
    rec = recall_score(y_test, y_pred, zero_division=0)
    f1 = f1_score(y_test, y_pred, zero_division=0)
    try:
        auc = roc_auc_score(y_test, clf.decision_function(X_test_scaled))
    except ValueError:
        auc = 0.0

    class_count = int(np.min(np.bincount(y)))
    folds = min(5, class_count)
    cv_mean = None
    cv_std = None
    if folds >= 2:
        full_scaler = StandardScaler()
        X_scaled = full_scaler.fit_transform(X)
        cv = StratifiedKFold(folds, shuffle=True, random_state=RANDOM_STATE)
        scores = cross_val_score(clf, X_scaled, y, cv=cv)
        cv_mean = float(np.mean(scores))
        cv_std = float(np.std(scores))

    return {
        'accuracy': float(acc),
        'precision': float(prec),
        'recall': float(rec),
        'f1': float(f1),
        'auc': float(auc),
        'cv_mean': cv_mean,
        'cv_std': cv_std,
        'segments_per_class': class_count,
        'confusion_matrix': confusion_matrix(y_test, y_pred),
    }


# ==============================================================================
# TRAIN & EVALUATE
# ==============================================================================

def run_detector_experiment(cover_path=None):
    """Run the detector experiment: train SVM, evaluate, print report."""
    print("=" * 70)
    print("   STEGANALYSIS DETECTOR-BASED EVALUATION (ML Benchmark)")
    print(f"   Audio: {cover_path or COVER_PATH}")
    print("=" * 70)

    print("\n[1] Building dataset...")
    X, y, cover_path, stego_paths = build_dataset(cover_path)
    n_cover = np.sum(y == 0)
    n_stego = np.sum(y == 1)
    print(f"    Cover segments : {n_cover}")
    print(f"    Stego segments : {n_stego}")
    print(f"    Total samples  : {len(y)}")
    print(f"    Feature dim    : {X.shape[1]}")
    print(f"    Stego files    : {len(stego_paths)}")

    # Handle class imbalance: stratify
    if n_cover == 0 or n_stego == 0:
        print("\n[ERROR] Not enough data to train/evaluate the detector. Please ensure both cover and stego segments are present.")
        return

    print("\n[2] Split train/test (stratified)...")
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y
    )
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)

    print("\n[3] Training SVM (RBF kernel)...")
    clf = SVC(kernel='rbf', C=1.0, gamma='scale', random_state=RANDOM_STATE)
    clf.fit(X_train_scaled, y_train)
    y_pred = clf.predict(X_test_scaled)

    print("\n[4] Evaluation...")
    acc = accuracy_score(y_test, y_pred)
    prec = precision_score(y_test, y_pred, zero_division=0)
    rec = recall_score(y_test, y_pred, zero_division=0)
    f1 = f1_score(y_test, y_pred, zero_division=0)

    try:
        y_prob = clf.decision_function(X_test_scaled)
        auc = roc_auc_score(y_test, y_prob)
    except Exception:
        auc = 0.0

    cm = confusion_matrix(y_test, y_pred)

    # Cross-validation
    print("\n[5] Cross-validation (5-fold stratified)...")
    X_scaled = scaler.fit_transform(X)
    cv_scores = cross_val_score(clf, X_scaled, y, cv=StratifiedKFold(5, shuffle=True, random_state=RANDOM_STATE))
    cv_mean = np.mean(cv_scores)
    cv_std = np.std(cv_scores)

    # --- REPORT ---
    print("\n" + "=" * 70)
    print("   RESULTS OF DETECTOR-BASED STEGANALYSIS EXPERIMENT")
    print("=" * 70)
    print(f"\n  Confusion Matrix (Test Set):")
    print(f"                   Predicted")
    print(f"                 Cover  Stego")
    print(f"    Actual Cover   {cm[0,0]:>4}    {cm[0,1]:>4}")
    print(f"    Actual Stego   {cm[1,0]:>4}    {cm[1,1]:>4}")
    print(f"\n  Metrics (Test Set):")
    print(f"    Accuracy       : {acc*100:.2f}%")
    print(f"    Precision      : {prec*100:.2f}%")
    print(f"    Recall         : {rec*100:.2f}%")
    print(f"    F1-Score       : {f1*100:.2f}%")
    print(f"    AUC-ROC        : {auc*100:.2f}%")
    print(f"\n  Cross-Validation (5-fold):")
    print(f"    Mean Accuracy  : {cv_mean*100:.2f}% (+/- {cv_std*100:.2f}%)")
    print("\n" + "-" * 70)

    return {
        'accuracy': acc,
        'precision': prec,
        'recall': rec,
        'f1': f1,
        'auc': auc,
        'cv_mean': cv_mean,
        'cv_std': cv_std,
        'confusion_matrix': cm,
    }


# ==============================================================================
# MAIN
# ==============================================================================


"""
Compact Steganalysis: NC and Entropy.
Focus on audio-quality metrics and Excel export.

Metrics:
    1. Entropy           - Entropy difference between cover and stego (smaller is better)
    2. NC                - Normalized Correlation cover vs stego (larger is better, ideal = 1.0)

Usage:
        python steganalysis.py
"""

import numpy as np
import scipy.io.wavfile as scp
from collections import Counter
import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import math


# ==============================================================================
# UTILITY
# ==============================================================================

def load_audio_samples(filepath):
    rate, data = scp.read(filepath)
    if data.ndim > 1:
        data = data[:, 0]
    return rate, data.astype(np.float64)


def _build_interpolated_reference(cover_samples):
    """
    Reconstruct a reference signal aligned with stego audio.
    Stego audio inserts interpolated samples between original samples,
    so its length is ~2N-1. This function creates a clean (non-embedded)
    reference for fair metric comparison.
    """
    n = len(cover_samples)
    ref = np.zeros(2 * n - 1)
    ref[::2] = cover_samples
    ref[1::2] = np.floor((cover_samples[:-1] + cover_samples[1:]) / 2.0)
    return ref


# ==============================================================================
# 1. ENTROPY
# ==============================================================================

def calculate_entropy(samples):
    """
    Shannon Entropy: H(X) = -sum(p(x) * log2(p(x)))
    A smaller cover-vs-stego entropy gap is better.
    Ideal value: difference = 0 (identical distribution).
    """
    samples_int = samples.astype(np.int64)
    counts = Counter(samples_int)
    total = len(samples_int)
    entropy = 0.0
    for count in counts.values():
        p = count / total
        if p > 0:
            entropy -= p * math.log2(p)
    return entropy


def entropy_analysis(cover_samples, stego_samples):
    if len(stego_samples) >= len(cover_samples) * 1.8:
        ref = _build_interpolated_reference(cover_samples)
    else:
        ref = cover_samples
    e_cover = calculate_entropy(ref)
    e_stego = calculate_entropy(stego_samples)
    delta = abs(e_stego - e_cover)
    relative_change = (delta / e_cover * 100) if e_cover > 0 else 0
    return {
        'entropy_cover': e_cover,
        'entropy_stego': e_stego,
        'entropy_diff': delta,
        'entropy_pct': relative_change
    }


# ==============================================================================
# 2. NORMALIZED CORRELATION (NC)
# ==============================================================================

def normalized_correlation(cover_samples, stego_samples):
    """
    NC = sum(C*S) / sqrt(sum(C^2) * sum(S^2))
    Higher NC is better. Ideal value: NC = 1.0.

    Handles interpolated stego (length ~2x cover) by building
    an aligned reference signal.
    """
    if len(stego_samples) >= len(cover_samples) * 1.8:
        c = _build_interpolated_reference(cover_samples)
    else:
        c = cover_samples

    min_len = min(len(c), len(stego_samples))
    c = c[:min_len]
    s = stego_samples[:min_len]
    numerator = np.sum(c * s)
    denominator = np.sqrt(np.sum(c ** 2) * np.sum(s ** 2))
    nc = numerator / denominator if denominator > 0 else 0
    return nc


def evaluate_audio_pair(cover_path, stego_path):
    """Evaluate Entropy and NC for one selected cover/stego pair."""
    if not os.path.isfile(cover_path):
        raise FileNotFoundError(f"Cover audio not found: {cover_path}")
    if not os.path.isfile(stego_path):
        raise FileNotFoundError(f"Stego audio not found: {stego_path}")

    _, cover = load_audio_samples(cover_path)
    _, stego = load_audio_samples(stego_path)

    entropy = entropy_analysis(cover, stego)
    return {
        **entropy,
        'nc': normalized_correlation(cover, stego),
    }


# ==============================================================================
# DETECT SHARES
# ==============================================================================

def _detect_total_shares(audio_no, payload_no):
    idx = 0
    while os.path.exists(f'STEGOAUDIO/stego_audio{audio_no}_payload{payload_no}/stegoaudio{idx}.wav'):
        idx += 1
    return idx


# ==============================================================================
# EXCEL STYLING
# ==============================================================================

def _style_header(ws, row, max_col):
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border


def _dc(cell, fmt=None):
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt


PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ==============================================================================
# EXPORT EXCEL
# ==============================================================================

def export_to_excel(all_rows, all_aggregate, excel_path):
    """
    Export to Excel with sheets:
      1. Share Details              : Audio, Payload, Share, Entropy, NC
      2. Summary (Avg)              : Average per Audio-Payload
      3. Summary (Best)             : Best value per Audio-Payload
      4. Summary (Worst)            : Worst value per Audio-Payload
      5+. Pivot per metric per share: Audio x Payload matrix (chart-ready)
      Last. Metric Notes            : Theory & thresholds
    """
    wb = xl.Workbook()
    wb.remove(wb.active)

    # ----- Sheet 1: Share Details -----
    ws1 = wb.create_sheet("Share Details")
    headers = [
        "Audio", "Payload", "Share",
        "Entropy Cover (bits)", "Entropy Stego (bits)",
        "Entropy Diff (bits)", "Entropy Change (%)",
        "NC"
    ]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    _style_header(ws1, 1, len(headers))

    for r, rd in enumerate(all_rows, 2):
        m = rd['m']
        vals = [
            rd['audio'], rd['payload'], rd['share'],
            m['entropy_cover'], m['entropy_stego'],
            m['entropy_diff'], m['entropy_pct'],
            m['nc']
        ]
        fmts = [
            None, None, None,
            '0.000000', '0.000000',
            '0.000000', '0.0000',
            '0.0000000000'
        ]
        for c, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws1.cell(row=r, column=c, value=v)
            _dc(cell, f)

    for c in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 20

    # ----- Sheets 2-4: Summary Avg / Best / Worst -----
    summary_cols = [
        ("Entropy Change (%)", 'entropy_pct', '0.0000', 'min'),
        ("NC", 'nc', '0.0000000000', 'max'),
    ]

    for label, agg_type in [("Summary (Avg)", 'avg'), ("Summary (Best)", 'best'), ("Summary (Worst)", 'worst')]:
        ws = wb.create_sheet(label)
        sh = ["Audio", "Payload", "#Share"] + [s[0] for s in summary_cols]
        for c, h in enumerate(sh, 1):
            ws.cell(row=1, column=c, value=h)
        _style_header(ws, 1, len(sh))

        for r, agg in enumerate(all_aggregate, 2):
            ws.cell(row=r, column=1, value=agg['audio']); _dc(ws.cell(row=r, column=1))
            ws.cell(row=r, column=2, value=agg['payload']); _dc(ws.cell(row=r, column=2))
            ws.cell(row=r, column=3, value=agg['n_shares']); _dc(ws.cell(row=r, column=3))

            for ci, (_, key, fmt, direction) in enumerate(summary_cols):
                vals = [m[key] for m in agg['metrics']]
                if agg_type == 'avg':
                    val = np.mean(vals)
                elif agg_type == 'best':
                    val = max(vals) if direction == 'max' else min(vals)
                else:
                    val = min(vals) if direction == 'max' else max(vals)
                cell = ws.cell(row=r, column=4 + ci, value=val)
                _dc(cell, fmt)

        for c in range(1, len(sh) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 20

    # ----- Pivot sheets per metric per share -----
    all_shares = sorted(set(r['share'] for r in all_rows))
    all_audios = sorted(set(r['audio'] for r in all_rows))
    all_payloads = sorted(set(r['payload'] for r in all_rows))
    lookup = {(r['audio'], r['payload'], r['share']): r['m'] for r in all_rows}

    pivot_defs = [
        ("Entropy (%)", 'entropy_pct', '0.0000', 1, 'min'),
        ("NC", 'nc', '0.0000000000', 0.999, 'max'),
    ]

    for metric_label, key, fmt, threshold, direction in pivot_defs:
        for s in all_shares:
            sname = f"{metric_label} S{s}"
            if len(sname) > 31:
                sname = sname[:31]
            ws = wb.create_sheet(sname)

            ws.cell(row=1, column=1, value=f"{metric_label} (Share {s})")
            ws.cell(row=1, column=1).font = Font(bold=True, size=12)

            ws.cell(row=2, column=1, value="Audio \\ Payload")
            for ci, p in enumerate(all_payloads):
                ws.cell(row=2, column=ci + 2, value=f"Payload {p}")
            _style_header(ws, 2, len(all_payloads) + 1)

            for ri, a in enumerate(all_audios):
                lbl = ws.cell(row=ri + 3, column=1, value=f"Audio {a}")
                lbl.font = Font(bold=True)
                _dc(lbl)

                for ci, p in enumerate(all_payloads):
                    m = lookup.get((a, p, s))
                    if m:
                        val = m[key]
                        cell = ws.cell(row=ri + 3, column=ci + 2, value=val)
                        _dc(cell, fmt)
                        if direction == 'min':
                            cell.fill = PASS_FILL if val < threshold else FAIL_FILL
                        else:
                            cell.fill = PASS_FILL if val >= threshold else FAIL_FILL
                    else:
                        cell = ws.cell(row=ri + 3, column=ci + 2, value="-")
                        _dc(cell)

            ws.column_dimensions['A'].width = 14
            for ci in range(len(all_payloads)):
                ws.column_dimensions[get_column_letter(ci + 2)].width = 14

    # ----- Metric Notes sheet -----
    ws_k = wb.create_sheet("Metric Notes")
    info = [
        ["Metric", "Preferred Direction", "PASS Threshold", "Ideal Value", "Formula / Notes"],
        ["Entropy Change (%)",
         "Lower is better", "< 1%", "0%",
         "Relative difference of Shannon entropy between cover and stego. H(X) = -sum(p(x)*log2(p(x)))"],
        ["NC (Normalized Correlation)",
         "Higher is better", ">= 0.999", "1.0",
         "Signal correlation. NC = sum(C*S) / sqrt(sum(C^2)*sum(S^2)). Range: -1 to 1"],
    ]
    for ri, row in enumerate(info):
        for ci, val in enumerate(row):
            cell = ws_k.cell(row=ri + 1, column=ci + 1, value=val)
            _dc(cell)
    _style_header(ws_k, 1, 5)
    for c in range(1, 6):
        ws_k.column_dimensions[get_column_letter(c)].width = 32

    os.makedirs(os.path.dirname(excel_path) if os.path.dirname(excel_path) else '.', exist_ok=True)
    wb.save(excel_path)
    print(f"\n  [OK] Results saved to: {excel_path}")


# ==============================================================================
# SINGLE ANALYSIS
# ==============================================================================

def run_single(audio_no='1', payload_no='1', stego_share=0, cover_path=None):
    cover_path = cover_path or f'DATASET/Audio/data{audio_no}_mono.wav'
    stego_path = f'STEGOAUDIO/stego_audio{audio_no}_payload{payload_no}/stegoaudio{stego_share}.wav'

    print("=" * 70)
    print("   STEGANALYSIS (NC and Entropy)")
    print(f"   Cover : {cover_path}")
    print(f"   Stego : {stego_path}")
    print("=" * 70)

    if not os.path.exists(cover_path):
        print(f"[ERROR] Cover audio not found: {cover_path}"); return
    if not os.path.exists(stego_path):
        print(f"[ERROR] Stego audio not found: {stego_path}"); return

    _, cover = load_audio_samples(cover_path)
    _, stego = load_audio_samples(stego_path)

    # --- Entropy ---
    print("\n" + "=" * 70)
    print("  1. ENTROPY ANALYSIS")
    ent = entropy_analysis(cover, stego)
    print(f"  Entropy cover  : {ent['entropy_cover']:.6f} bits")
    print(f"  Entropy stego  : {ent['entropy_stego']:.6f} bits")
    print(f"  Difference     : {ent['entropy_diff']:.6f} bits")
    print(f"  Change         : {ent['entropy_pct']:.4f}%")

    # --- NC ---
    print("\n" + "=" * 70)
    print("  2. NORMALIZED CORRELATION (NC)")
    nc = normalized_correlation(cover, stego)
    print(f"  NC             : {nc:.10f}")

    print("=" * 70)


# ==============================================================================
# BATCH ANALYSIS + EXCEL
# ==============================================================================

def run_batch(total_audio=15, total_payload=11, total_shares=None,
              excel_output=None):
    all_rows = []
    all_aggregate = []

    print("=" * 100)
    print("   BATCH STEGANALYSIS (NC and Entropy) - all shares")
    print("=" * 100)

    header = (f"{'Audio':>6} | {'Payload':>7} | {'Share':>5} | "
              f"{'Entropy%':>9} | {'NC':>14}")
    print(header)
    print("-" * 60)

    for x in range(1, total_audio + 1):
        for y in range(1, total_payload + 1):
            cover_path = f'DATASET/Audio/data{x}_mono.wav'
            if not os.path.exists(cover_path):
                continue

            n_shares = total_shares if total_shares else _detect_total_shares(x, y)
            if n_shares == 0:
                continue

            _, cover = load_audio_samples(cover_path)

            combo_metrics = []

            for s in range(n_shares):
                stego_path = f'STEGOAUDIO/stego_audio{x}_payload{y}/stegoaudio{s}.wav'
                if not os.path.exists(stego_path):
                    continue

                _, stego = load_audio_samples(stego_path)
                ent = entropy_analysis(cover, stego)
                nc = normalized_correlation(cover, stego)

                m = {
                    'entropy_cover': ent['entropy_cover'],
                    'entropy_stego': ent['entropy_stego'],
                    'entropy_diff': ent['entropy_diff'],
                    'entropy_pct': ent['entropy_pct'],
                    'nc': nc,
                }
                combo_metrics.append(m)
                all_rows.append({'audio': x, 'payload': y, 'share': s, 'm': m})

                row = (f"{x:>6} | {y:>7} | {s:>5} | "
                       f"{m['entropy_pct']:>8.4f}% | "
                       f"{nc:>14.10f}")
                print(row)

            if combo_metrics:
                all_aggregate.append({
                    'audio': x, 'payload': y,
                    'n_shares': len(combo_metrics),
                    'metrics': combo_metrics
                })

    print("=" * 100)

    if not all_aggregate:
        print("\n[INFO] No stego audio data was found.")
        return

    # --- Summary ---
    print("\n" + "=" * 100)
    print("   SUMMARY PER COMBINATION (avg / best / worst)")
    print("   Entropy%: best=min, worst=max | NC: best=max, worst=min")
    print("=" * 100)

    header2 = (f"{'Audio':>6} | {'Payload':>7} | {'#Sh':>4} | "
               f"{'Ent% avg':>9} | {'Ent% best':>9} | {'Ent% worst':>10} | "
               f"{'NC avg':>14} | {'NC best':>14} | {'NC worst':>14}")
    print(header2)
    print("-" * 100)

    for agg in all_aggregate:
        ms = agg['metrics']
        ent_vals = [m['entropy_pct'] for m in ms]
        nc_vals = [m['nc'] for m in ms]
        row2 = (f"{agg['audio']:>6} | {agg['payload']:>7} | {agg['n_shares']:>4} | "
                f"{np.mean(ent_vals):>8.4f}% | {min(ent_vals):>8.4f}% | {max(ent_vals):>9.4f}% | "
                f"{np.mean(nc_vals):>14.10f} | {max(nc_vals):>14.10f} | {min(nc_vals):>14.10f}")
        print(row2)

    print("=" * 100)

    # --- Export Excel ---
    if excel_output is None:
        n_str = str(total_shares) if total_shares else 'auto'
        excel_output = f'steganalysis_results/steganalysis_{n_str}_shares.xlsx'

    export_to_excel(all_rows, all_aggregate, excel_output)
    return all_rows, all_aggregate

def main():
    print("=" * 70)
    print("  STEGANALYSIS TOOLKIT")
    print("=" * 70)
    print("1. Detector ML")
    print("2. NC and Entropy Analysis (single)")
    print("3. NC and Entropy Analysis (batch + Excel)")
    choice = input("Select mode (1-3): ").strip()

    if choice == '1':
        print("=" * 70)
        print("  DETECTOR-BASED STEGANALYSIS (ML Benchmark)")
        print("=" * 70)

        print("Select stego_audio[X]_payload[Y] to analyze:")
        audio_no = input("Enter audio number [X]: ").strip() or '1'
        payload_no = input("Enter payload number [Y]: ").strip() or '1'
        if not audio_no.isdigit() or not payload_no.isdigit():
            raise ValueError("Audio number and payload number must be integers.")
        cover_path = build_cover_path(audio_no)
        run_detector_experiment(cover_path)

    elif choice == '2':
        print("=" * 70)
        print("  NC AND ENTROPY ANALYSIS (single pair)")
        print("=" * 70)

        print("Select stego_audio[X]_payload[Y] to analyze:")
        audio_no = input("Enter audio number [X]: ").strip() or '1'
        payload_no = input("Enter payload number [Y]: ").strip() or '1'
        if not audio_no.isdigit() or not payload_no.isdigit():
            raise ValueError("Audio number and payload number must be integers.")
        cover_path = build_cover_path(audio_no)
        share_input = input("Stego share number (default=0): ").strip()
        share_no = int(share_input) if share_input else 0
        run_single(audio_no, payload_no, share_no, cover_path)

    elif choice == '3':
        print("=" * 70)
        print("  NC AND ENTROPY ANALYSIS (batch + Excel)")
        print("=" * 70)

        audio_input = input("Total audio files (default=15): ").strip()
        payload_input = input("Total payloads (default=11): ").strip()
        shares_input = input("Total shares / n (blank = auto-detect): ").strip()
        run_batch(
            total_audio=int(audio_input) if audio_input else 15,
            total_payload=int(payload_input) if payload_input else 11,
            total_shares=int(shares_input) if shares_input else None,
        )
    else:
        print("Invalid selection.")


if __name__ == '__main__':
    main()
