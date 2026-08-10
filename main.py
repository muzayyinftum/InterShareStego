import argparse
import math
import os
import random
import sys
import time
import zipfile
from collections import Counter
from datetime import datetime

import numpy as np
import scipy.io.wavfile as scp
from scipy.fft import fft
from scipy.stats import kurtosis, skew
from sympy import nextprime

try:
    import tracemalloc
except ImportError:
    tracemalloc = None


MENU = {
    "1": ("Embedding", None),
    "2": ("Extracting", None),
    "3": ("Evaluation", None),
    "4": ("Exit", None),
}

MENU_ORDER = ("1", "2", "3", "4")

EVALUATION_MENU = {
    "1": ("Compare payload and audio", "single_compare"),
    "2": ("Check audio quality", "quality_evaluation"),
    "3": ("Detector ML", "detector_ml"),
    "4": ("NC and entropy analysis (single)", "nc_entropy_single"),
    "5": ("NC and entropy analysis (batch + excel)", "nc_entropy_batch"),
}

EVALUATION_MENU_ORDER = ("1", "2", "3", "4", "5")


def read_payload(filepath):
    with open(filepath, 'rb') as file:
        raw_data = file.read()

    # UTF-16 dengan BOM
    if raw_data.startswith(b'\xff\xfe') or raw_data.startswith(b'\xfe\xff'):
        data = raw_data.decode('utf-16')

    # Tanpa BOM
    else:
        data = raw_data.decode('utf-8')

    # Hilangkan TAB, spasi, newline, dll.
    binary_data = ''.join(data.split())

    # Validasi hanya 0 dan 1
    if not all(bit in '01' for bit in binary_data):
        invalid = set(binary_data) - {'0', '1'}
        raise ValueError(
            f"Payload contains invalid characters: {invalid}"
        )

    return binary_data


def sampling(file_audio):
    rate, data = scp.read(file_audio)
    data = np.add(np.int16(data), [32768])
    return rate, data


def interpolation_linear(input_sampling):
    index_odd = [x for x in range(0, (len(input_sampling) * 2) - 1) if x % 2 == 1]
    index_even = [x for x in range(0, (len(input_sampling) * 2)) if x % 2 == 0]
    interpolated_sample = np.interp(index_odd, index_even, input_sampling)
    interpolated_sample = np.floor(interpolated_sample)
    return interpolated_sample


def determine_sample_space(interpolated_sample):
    bit = []
    for x in range(len(interpolated_sample)):
        if interpolated_sample[x] == 0:
            bit.append(0)
        else:
            bit.append(math.floor(math.sqrt(math.log(interpolated_sample[x], 2))))
    return bit


def segmentation(payload, bit):
    index = 0
    processed_payload = []
    for x in bit:
        if index >= len(payload):
            break
        processed_payload.append(payload[index:index + x])
        index += x

    return processed_payload, len(processed_payload[-1])


def convert_bin_to_dec(payload):
    decimal = []
    isZeroInLast = False
    for x in range(len(payload)):
        decimal.append(int(payload[x], 2))

    if decimal[-1] == 0:
        isZeroInLast = True
    return decimal, isZeroInLast


def get_prime_number(decimal_payload):
    max_value = max(decimal_payload)
    prime = nextprime(max_value)
    return prime


def shamir_secret_sharing(decimal_payload, prime, total_shares, min_shares):
    all_data_shares = []
    for x in range(len(decimal_payload)):
        data = split_secret(decimal_payload[x], prime, total_shares, min_shares)
        all_data_shares.append(data)

    return all_data_shares


def split_secret(secret, prime, total_shares, min_shares):
    coefficients = [secret]
    for _ in range(min_shares - 1):
        random_coefficient = random.randint(1, prime - 1)
        coefficients.append(random_coefficient)

    shares = []
    for m in range(1, total_shares + 1):
        y = evaluate_polynomial(coefficients, m, prime)
        shares.append(y)
    return shares


def evaluate_polynomial(coefficients, x, prime):
    result = 0
    for i, c in enumerate(coefficients):
        term = (c * (x ** i)) % prime
        result = (result + term) % prime
    return result


def embedding(data, interpolated_sample, total_shares, last_bit, isZeroInLast):
    all_data = []
    for i in range(total_shares):
        single_audio = []
        for x in range(len(interpolated_sample)):
            if x <= len(data) - 1:
                single_audio.append(interpolated_sample[x] - data[x][i])
            else:
                if x == len(interpolated_sample) - 1:
                    single_audio.append(interpolated_sample[x] - i - 1)
                elif x == len(interpolated_sample) - 2:
                    single_audio.append(interpolated_sample[x] - int(last_bit))
                elif x == len(interpolated_sample) - 3 and isZeroInLast:
                    single_audio.append(interpolated_sample[x] - 1)
                else:
                    single_audio.append(interpolated_sample[x])

        all_data.append(single_audio)
    return all_data


def combine(embedded, original_sample):
    all_data = []
    for x in range(len(embedded)):
        single_stego = []
        index_stego = 0
        index_sample = 0
        index_embed = 0
        for y in range(0, len(original_sample) * 2 - 1):
            if index_stego % 2 == 0:
                single_stego.append(original_sample[index_sample])
                index_sample += 1
            else:
                single_stego.append(embedded[x][index_embed])
                index_embed += 1
            index_stego += 1
        all_data.append(single_stego)
    return all_data


def create_stego_audio(stego_data, filepath, cover_sample_rate):
    for x in range(len(stego_data)):
        new_filepath = filepath + str(x) + '.wav'
        process_1 = np.subtract(stego_data[x], [32768])
        stego_audio = np.array(process_1, dtype=np.int16)
        os.makedirs(os.path.dirname(new_filepath), exist_ok=True)
        length = len(stego_audio)
        n_cover = (length + 1) // 2
        stego_sample_rate = int(round(cover_sample_rate * length / n_cover))
        scp.write(new_filepath, stego_sample_rate, stego_audio)


def extraction_sampling(stego_audio_file, stego_audio_no):
    all_stego_audio = []
    rate = None
    for x in stego_audio_no:
        new_filepath = stego_audio_file + str(x) + '.wav'
        rate, data = scp.read(new_filepath)
        data = np.add(np.int16(data), [32768])
        all_stego_audio.append(data)
    return rate, all_stego_audio


def separate(stego_sample):
    original_sample = []
    embedded = []
    for x in range(len(stego_sample)):
        single_original = []
        single_embedded = []
        for y in range(len(stego_sample[x])):
            if y % 2 == 0:
                single_original.append(stego_sample[x][y])
            else:
                single_embedded.append(stego_sample[x][y])
        original_sample.append(single_original)
        embedded.append(single_embedded)

    return original_sample, embedded


def extraction_interpolation_linear(embedded):
    all_interpolated = []
    for x in range(len(embedded)):
        index_odd = [x for x in range(0, (len(embedded[x]) * 2) - 1) if x % 2 == 1]
        index_even = [x for x in range(0, (len(embedded[x]) * 2)) if x % 2 == 0]
        interpolated_sample = np.interp(index_odd, index_even, embedded[x])
        interpolated_sample = np.floor(interpolated_sample)
        all_interpolated.append(interpolated_sample)
    return all_interpolated


def check_last_index(embedded, interpolated_sample):
    last_index_of_embedding = []
    tmpZeroLast = []
    for x in range(len(embedded)):
        single_last_decimal = []
        single_zero_last = []
        for y in range(len(embedded[x]) - 1, -1, -1):
            value = int(interpolated_sample[x][y] - embedded[x][y])

            if value != 0 and y == len(embedded[x]) - 3:
                single_zero_last.append(value)
            if value != 0 and y != len(embedded[x]) - 1 and y != len(embedded[x]) - 2 and y != len(embedded[x]) - 3:
                single_last_decimal.append(y)
                break
        last_index_of_embedding.append(single_last_decimal)
        if len(single_zero_last) > 0:
            tmpZeroLast.append(single_zero_last)

    last_index = last_index_of_embedding[0][0]
    if len(last_index_of_embedding) > 1:
        value = all(sublist[0] == last_index for sublist in last_index_of_embedding)
        if value is False:
            flattened_list = [item for sublist in last_index_of_embedding for item in sublist]
            last_index = max(flattened_list)
    else:
        if len(tmpZeroLast) > 0:
            last_index += 1
    return last_index, last_index_of_embedding


def check_next_prime(difference_data):
    all_prime = []
    for x in range(len(difference_data)):
        prime = get_prime_number(difference_data[x])
        all_prime.append(prime)
    first_value = all_prime[0]
    value = all(x == first_value for x in all_prime)
    return value, first_value


def extraction_difference_determination(embedded, interpolated_sample):
    last_index, all_last_index = check_last_index(embedded, interpolated_sample)

    difference_data = []
    data_sahre_no = []
    last_bit = []
    for x in range(len(embedded)):
        single_decimal = []
        for y in range(len(embedded[x])):
            if y <= last_index:
                value = int(interpolated_sample[x][y] - embedded[x][y])
                single_decimal.append(value)
            if y == len(embedded[x]) - 2:
                single_last_bit = int(interpolated_sample[x][y] - embedded[x][y])
                last_bit.append(single_last_bit)
            if y == len(embedded[x]) - 1:
                single_share_no = int(interpolated_sample[x][y] - embedded[x][y])
                data_sahre_no.append(single_share_no)

        difference_data.append(single_decimal)

    prime_value, next_prime = check_next_prime(difference_data)

    if prime_value is False:
        print("The prime values are not the same:", next_prime)
        return

    return difference_data, last_index + 1, next_prime, data_sahre_no, last_bit[0]


def extraction_determine_sample_space(interpolated_sample, last_index):
    bit = []
    for x in range(len(interpolated_sample)):
        single_audio_bit = []
        for y in range(len(interpolated_sample[x])):
            if y == last_index:
                break
            single_audio_bit.append(math.floor(math.sqrt(math.log(interpolated_sample[x][y], 2))))
        bit.append(single_audio_bit)
    return bit


def transpose_matrix(matrix):
    return [[matrix[j][i] for j in range(len(matrix))] for i in range(len(matrix[0]))]


def reconstruct_secret(difference_data, prime_number, share_no):
    transposed_difference = transpose_matrix(difference_data)

    all_data = []
    for x in range(len(transposed_difference)):
        combined_list = list(zip(share_no, transposed_difference[x]))
        data = reconstruct_secret2(combined_list, prime_number)
        all_data.append(data)
    return all_data


def reconstruct_secret2(shares, prime):
    def lagrange_interpolation(x, x_s, y_s, prime):
        total = 0
        for i in range(len(x_s)):
            xi, yi = x_s[i], y_s[i]
            prod = yi
            for j in range(len(x_s)):
                if i != j:
                    xj = x_s[j]
                    prod *= (x - xj) * pow(xi - xj, -1, prime)
                    prod %= prime
            total += prod
            total %= prime
        return total

    x_s, y_s = zip(*shares)
    return lagrange_interpolation(0, x_s, y_s, prime)


def check_all_nested_arrays_equal(nested_arr):
    def check_all_equal(arr):
        first_value = arr[0]
        for x in arr:
            if x != first_value:
                return False
        return True

    for sub_arr in nested_arr:
        if not check_all_equal(sub_arr):
            return False
    return True


def decimal_to_binary(decimal_payload, bit, last_bit):
    binary_payload = []
    new_bit = bit[0]
    for x in range(len(decimal_payload)):
        if x == len(decimal_payload) - 1:
            binary_payload.append(np.binary_repr(decimal_payload[x], width=last_bit))
        else:
            binary_payload.append(np.binary_repr(decimal_payload[x], width=new_bit[x]))
    translated_payload = ''.join(binary_payload)

    return translated_payload


def create_payload(byte_payload, filepath):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, 'w+') as file:
        file.write(byte_payload)


def create_cover_audio(original_sample, filepath):
    unnormalize_data = np.subtract(original_sample, [32768])
    new_data_sample = np.array(unnormalize_data, dtype=np.int16)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    scp.write(filepath, 44100, new_data_sample)


def calculate_ber(original_binary, extracted_binary):
    def to_binary_str(data):
        if isinstance(data, (list, tuple)):
            return ''.join(str(b) for b in data if str(b) in '01')
        content = str(data).replace(' ', '').replace('\t', '')
        return ''.join(c for c in content if c in '01')

    original_binary = to_binary_str(original_binary)
    extracted_binary = to_binary_str(extracted_binary)

    min_length = min(len(original_binary), len(extracted_binary))

    if min_length == 0:
        return 100.0, 0, 0

    bit_errors = 0
    for i in range(min_length):
        if original_binary[i] != extracted_binary[i]:
            bit_errors += 1

    total_bits = min_length
    ber_percentage = (bit_errors / total_bits) * 100

    return ber_percentage, bit_errors, total_bits


def show_menu():
    print("\n" + "=" * 70)
    print("  STEGOSHARE - Reversible Audio Steganography Using Secret Sharing")
    print("=" * 70)

    for choice in MENU_ORDER:
        name, _ = MENU[choice]
        print("{}. {}".format(choice, name))


def show_menu_evaluation():
    print("\n" + "=" * 70)
    print("  EVALUATION")
    print("=" * 70)

    for choice in EVALUATION_MENU_ORDER:
        name, _ = EVALUATION_MENU[choice]
        print("{}. {}".format(choice, name))


def build_output_base(audio_file, payload_file, output_root='results/STEGOAUDIO'):
    audio_name = os.path.splitext(os.path.basename(audio_file))[0]
    payload_name = os.path.splitext(os.path.basename(payload_file))[0]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = os.path.join(output_root, '{}_{}_{}'.format(audio_name, payload_name, timestamp))
    return os.path.join(output_dir, 'stegoaudio')


def run_single_embedding(payload_file, audio_file, total_shares, min_shares, output_base=None):
    if output_base is None:
        output_base = build_output_base(audio_file, payload_file)

    start = time.time()
    if tracemalloc is not None:
        tracemalloc.start()

    binary_payload = read_payload(payload_file)

    frame_rate, original_sample = sampling(audio_file)
    interpolated_sample = interpolation_linear(original_sample)

    bit = determine_sample_space(interpolated_sample)
    segmented_payload, last_bit = segmentation(binary_payload, bit)
    decimal_payload, isZeroInLast = convert_bin_to_dec(segmented_payload)
    prime_number = get_prime_number(decimal_payload)

    data_shares = shamir_secret_sharing(decimal_payload, prime_number, total_shares, min_shares)
    embedded = embedding(data_shares, interpolated_sample, total_shares, last_bit, isZeroInLast)

    stego_data = combine(embedded, original_sample)
    create_stego_audio(stego_data, output_base, frame_rate)

    if tracemalloc is not None:
        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
    else:
        peak = 0
    runtime = time.time() - start

    output_files = [
        '{}{}.wav'.format(output_base, index)
        for index in range(total_shares)
    ]

    return {
        'output_base': output_base,
        'output_dir': os.path.dirname(output_base),
        'output_files': output_files,
        'runtime': runtime,
        'peak_memory_mb': peak / 10**6,
        'total_shares': total_shares,
        'min_shares': min_shares,
    }


def run_embedding():
    audio_no = input("Enter audio number from DATASET: ").strip()
    payload_no = input("Enter payload number from DATASET: ").strip()

    if not audio_no.isdigit() or not payload_no.isdigit():
        raise ValueError("Audio number and payload number must be integers.")

    audio_file = 'DATASET/Audio/data{}_mono.wav'.format(audio_no)
    payload_file = 'DATASET/Payload/payload{}.txt'.format(payload_no)
    output_base = 'results/STEGOAUDIO/stego_audio{}_payload{}/stegoaudio'.format(
        audio_no,
        payload_no
    )

    if not os.path.isfile(audio_file):
        raise FileNotFoundError("Audio file not found: {}".format(audio_file))
    if not os.path.isfile(payload_file):
        raise FileNotFoundError("Payload file not found: {}".format(payload_file))

    total_shares, min_shares = map(
        int,
        input("Enter total and minimum shares (n k): ").split()
    )
    if total_shares < 1 or not 1 <= min_shares <= total_shares:
        raise ValueError("Value of shares must satisfy 1 <= k <= n.")

    result = run_single_embedding(
        payload_file,
        audio_file,
        total_shares,
        min_shares,
        output_base
    )

    print("\nEmbedding completed successfully.")
    print("Output folder:", result['output_dir'])
    print("Peak memory:", result['peak_memory_mb'], "MB")
    print("Embedding runtime:", result['runtime'])


def build_extraction_output_dir(zip_file, output_root='results/EXTRACTED'):
    zip_name = os.path.splitext(os.path.basename(zip_file))[0]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return os.path.join(output_root, '{}_{}'.format(zip_name, timestamp))


def extract_wav_files_from_zip(zip_file, output_dir):
    if not zipfile.is_zipfile(zip_file):
        raise ValueError("File input must be a ZIP file.")

    input_dir = os.path.join(output_dir, 'input_stego_audio')
    os.makedirs(input_dir, exist_ok=True)

    wav_files = []
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        for member in zip_ref.infolist():
            if member.is_dir() or not member.filename.lower().endswith('.wav'):
                continue

            filename = os.path.basename(member.filename)
            if not filename:
                continue

            target_path = os.path.join(input_dir, filename)
            with zip_ref.open(member) as source, open(target_path, 'wb') as target:
                target.write(source.read())
            wav_files.append(target_path)

    wav_files.sort()
    return wav_files


def read_stego_samples(stego_files):
    all_stego_audio = []
    sample_rate = None

    for stego_file in stego_files:
        rate, data = scp.read(stego_file)
        if sample_rate is None:
            sample_rate = rate

        data = np.add(np.int16(data), [32768])
        all_stego_audio.append(data)

    return sample_rate, all_stego_audio


def extract_payload_and_audio(stego_sample, output_dir):
    original_sample, embedded = separate(stego_sample)
    interpolated_sample = extraction_interpolation_linear(original_sample)

    extraction_result = extraction_difference_determination(embedded, interpolated_sample)
    if extraction_result is None:
        raise ValueError("Extraction failed because the prime values in the stego-audio were inconsistent.")

    difference_data, last_index, prime_number, share_no, last_bit = extraction_result
    bit = extraction_determine_sample_space(interpolated_sample, last_index)
    decimal_payload = reconstruct_secret(difference_data, prime_number, share_no)
    binary_payload = decimal_to_binary(decimal_payload, bit, last_bit)

    payload_output = os.path.join(output_dir, 'payload.txt')
    cover_audio_output = os.path.join(output_dir, 'audio.wav')
    create_payload(binary_payload, payload_output)
    create_cover_audio(original_sample[0], cover_audio_output)

    return payload_output, cover_audio_output


def run_single_extraction(zip_file, min_shares, output_dir=None):
    if output_dir is None:
        output_dir = build_extraction_output_dir(zip_file)

    start = time.time()
    os.makedirs(output_dir, exist_ok=True)

    wav_files = extract_wav_files_from_zip(zip_file, output_dir)
    if len(wav_files) < min_shares:
        raise ValueError(
            "ZIP file must contain at least {} WAV stego-audio files. Found {} files.".format(
                min_shares,
                len(wav_files)
            )
        )

    selected_stego_files = wav_files[:min_shares]
    frame_rate, stego_sample = read_stego_samples(selected_stego_files)
    payload_output, cover_audio_output = extract_payload_and_audio(stego_sample, output_dir)

    return {
        'output_dir': output_dir,
        'payload_file': payload_output,
        'cover_audio_file': cover_audio_output,
        'selected_stego_files': selected_stego_files,
        'zip_wav_count': len(wav_files),
        'min_shares': min_shares,
        'runtime': time.time() - start,
    }


def run_extraction():
    print("Select stego_audio[X]_payload[Y] to extract:")

    audio_no = input("Enter audio number [X]: ").strip()
    payload_no = input("Enter payload number [Y]: ").strip()

    if not audio_no.isdigit() or not payload_no.isdigit():
        raise ValueError("Audio number and payload number must be integers.")

    stego_audio_base = 'results/STEGOAUDIO/stego_audio{}_payload{}/stegoaudio'.format(
        audio_no,
        payload_no
    )
    output_dir = 'results/EXTRACTED/stego_audio{}_payload{}'.format(audio_no, payload_no)

    total_shares, min_shares = map(
        int,
        input("Enter total and minimum shares (n k): ").split()
    )
    if total_shares < 1 or not 1 <= min_shares <= total_shares:
        raise ValueError("Value of shares must satisfy 1 <= k <= n.")

    missing_files = [
        '{}{}.wav'.format(stego_audio_base, index)
        for index in range(total_shares)
        if not os.path.isfile('{}{}.wav'.format(stego_audio_base, index))
    ]
    if missing_files:
        raise FileNotFoundError(
            "Incomplete stego-audio file. First file not found: {}".format(
                missing_files[0]
            )
        )

    stego_audio_no = random.sample(range(total_shares), min_shares)

    frame_rate, stego_sample = extraction_sampling(stego_audio_base, stego_audio_no)
    payload_output, cover_audio_output = extract_payload_and_audio(stego_sample, output_dir)

    print("\nExtraction completed successfully.")
    print("Share used:", stego_audio_no)
    print("Extracted payload:", payload_output)
    print("Extracted audio:", cover_audio_output)


def sampling_quality(file_audio):
    rate, data = scp.read(file_audio)
    data = np.array(data, dtype=np.int16)
    data = np.add(data, [32768])
    return data


def mean_data_sample(data_sample):
    return np.mean(np.power(data_sample, [2]))


def calculate_mse(data_sample, data_stego):
    sample = np.asarray(data_sample, dtype=np.float64)
    stego = np.asarray(data_stego, dtype=np.float64)
    return np.mean(np.square(sample - stego))


def calculate_snr(data_sample, mse):
    if mse == 0:
        return 'infinite'

    mds = mean_data_sample(data_sample)
    log_content = mds / mse
    return 10 * math.log(log_content, 10)


def calculate_psnr(mse):
    if mse == 0:
        return 'infinite'

    log_content = (((2 ** 16) - 1) ** 2) / mse
    return 10 * math.log(log_content, 10)


def clone_cover_audio(data_sample, filename):
    index_odd = [x for x in range(0, (len(data_sample) * 2) - 1) if x % 2 == 1]
    index_even = [x for x in range(0, (len(data_sample) * 2)) if x % 2 == 0]

    interpolated_sample = np.interp(index_odd, index_even, data_sample)
    interpolated_sample = np.floor(interpolated_sample)

    new_data = []
    i_odd = 0
    i_even = 0
    for x in range(len(data_sample) * 2 - 1):
        if x % 2 == 0:
            new_data.append(data_sample[i_even])
            i_even += 1
        else:
            new_data.append(interpolated_sample[i_odd])
            i_odd += 1

    process_data = np.subtract(new_data, [32768])
    process_data = np.array(process_data, dtype=np.int16)
    output_dir = os.path.dirname(filename)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    scp.write(filename, 88200, process_data)
    return new_data


def print_excel(data_mse, data_snr, data_psnr, filename):
    import openpyxl as xl

    excel = xl.Workbook()
    sheet_mse = excel.create_sheet('Mean Squared Error')

    total_audio = len(data_mse)
    total_payload = len(data_mse[0])

    for index_audio in range(0, total_audio):
        if index_audio == 0:
            for x in range(0, total_payload):
                sheet_mse.cell(row=1, column=x + 2).value = 'Payload' + str(x + 1)

        for index_payload in range(0, total_payload):
            if index_payload == 0:
                sheet_mse.cell(row=index_audio + 2, column=1).value = 'Audio' + str(index_audio + 1)
            sheet_mse.cell(row=index_audio + 2, column=index_payload + 2).value = data_mse[index_audio][index_payload]

    sheet_snr = excel.create_sheet('SNR')

    total_audio = len(data_snr)
    total_payload = len(data_snr[0])

    for index_audio in range(0, total_audio):
        if index_audio == 0:
            for x in range(0, total_payload):
                sheet_snr.cell(row=1, column=x + 2).value = 'Payload' + str(x + 1)

        for index_payload in range(0, total_payload):
            if index_payload == 0:
                sheet_snr.cell(row=index_audio + 2, column=1).value = 'Audio' + str(index_audio + 1)
            sheet_snr.cell(row=index_audio + 2, column=index_payload + 2).value = data_snr[index_audio][index_payload]

    sheet_psnr = excel.create_sheet('PSNR')

    total_audio = len(data_psnr)
    total_payload = len(data_psnr[0])

    for index_audio in range(0, total_audio):
        if index_audio == 0:
            for x in range(0, total_payload):
                sheet_psnr.cell(row=1, column=x + 2).value = 'Payload' + str(x + 1)

        for index_payload in range(0, total_payload):
            if index_payload == 0:
                sheet_psnr.cell(row=index_audio + 2, column=1).value = 'Audio' + str(index_audio + 1)
            sheet_psnr.cell(row=index_audio + 2, column=index_payload + 2).value = data_psnr[index_audio][index_payload]

    excel.save(filename)


def run_single_quality(file_audio, file_stego_audio, clone_filename=None):
    if clone_filename is None:
        cover_name = os.path.splitext(os.path.basename(file_audio))[0]
        clone_filename = os.path.join('audio_clone', 'quality_clone_{}.wav'.format(cover_name))

    sample_audio = sampling_quality(file_audio)
    new_sample_audio = clone_cover_audio(sample_audio, clone_filename)

    sample_stego_audio = sampling_quality(file_stego_audio)

    if len(new_sample_audio) != len(sample_stego_audio):
        raise ValueError(
            "Length of clone audio ({}) and stego audio ({}) is not the same.".format(
                len(new_sample_audio),
                len(sample_stego_audio)
            )
        )

    mse = calculate_mse(new_sample_audio, sample_stego_audio)
    snr = calculate_snr(new_sample_audio, mse)
    psnr = calculate_psnr(mse)

    return {
        'mse': mse,
        'snr': snr,
        'psnr': psnr,
        'clone_file': clone_filename
    }


def run_quality_evaluation():
    print("-" * 70)
    print("  Quality Evaluation of Stego Audio")
    print("-" * 70)

    print("Select stego_audio[X]_payload[Y] to check quality:")
    audio = input("Enter audio number [X] (default=1): ").strip() or '1'
    payload = input("Enter payload number [Y] (default=1): ").strip() or '1'
    share = input("Enter stego audio number (default=0): ").strip() or '0'

    if not audio.isdigit() or not payload.isdigit() or not share.isdigit():
        raise ValueError("Audio number, payload number, and share number must be integers.")

    file_audio = 'DATASET/Audio/data{}_mono.wav'.format(audio)
    filename = 'results/CLONING/data_clone_audio{}.wav'.format(audio)
    file_stego_audio = 'results/STEGOAUDIO/stego_audio{}_payload{}/stegoaudio{}.wav'.format(
        audio,
        payload,
        share
    )

    if not os.path.isfile(file_audio):
        raise FileNotFoundError("Cover audio not found: {}".format(file_audio))
    if not os.path.isfile(file_stego_audio):
        raise FileNotFoundError("Stego audio not found: {}".format(file_stego_audio))

    result = run_single_quality(file_audio, file_stego_audio, filename)

    print("\nQuality evaluation from stego_audio{}_payload{}/stegoaudio{}.wav completed successfully.".format(
        audio,
        payload,
        share
    ))
    print("mse  :", result['mse'])
    print("snr  :", result['snr'])
    print("psnr :", result['psnr'])


def sampling_audio_for_compare(filepath):
    _, data = scp.read(filepath)
    data = np.asarray(data, dtype=np.int16)
    return np.add(data, [32768])


def compare_data(original_data, extracted_data, data_name):
    if len(original_data) != len(extracted_data):
        print("Original {} length : {}".format(data_name, len(original_data)))
        print("Extracted {} length: {}".format(data_name, len(extracted_data)))
        return False

    mismatch_indices = [
        index
        for index, (original, extracted) in enumerate(
            zip(original_data, extracted_data)
        )
        if not np.array_equal(original, extracted)
    ]

    if mismatch_indices:
        print("{} differences: {}".format(data_name.capitalize(), len(mismatch_indices)))
        print("First difference at index:", mismatch_indices[0])
        return False

    return True


def require_file(filepath):
    if not os.path.isfile(filepath):
        raise FileNotFoundError("File not found: {}".format(filepath))


def run_single_compare():
    print("-" * 70)
    print("  Compare original payload and audio with extracted results")
    print("-" * 70)

    print("Select stego_audio[X]_payload[Y] to extract:")
    audio_no = input("Enter audio number [X] (default=1): ").strip() or '1'
    payload_no = input("Enter payload number [Y] (default=1): ").strip() or '1'

    if not audio_no.isdigit() or not payload_no.isdigit():
        raise ValueError("Audio and payload numbers must be numeric.")

    original_payload = 'DATASET/Payload/payload{}.txt'.format(payload_no)
    extracted_payload = 'results/EXTRACTED/stego_audio{}_payload{}/payload.txt'.format(
        audio_no, payload_no
    )
    original_audio = 'DATASET/Audio/data{}_mono.wav'.format(audio_no)
    extracted_audio = 'results/EXTRACTED/stego_audio{}_payload{}/audio.wav'.format(
        audio_no, payload_no
    )

    for filepath in (
        original_payload, extracted_payload, original_audio, extracted_audio
    ):
        require_file(filepath)

    payload_equal = compare_data(
        read_payload(original_payload),
        read_payload(extracted_payload),
        'payload'
    )
    audio_equal = compare_data(
        sampling_audio_for_compare(original_audio),
        sampling_audio_for_compare(extracted_audio),
        'audio'
    )

    print("\nComparison results:")
    print("Payload:", "Exact" if payload_equal else "Different")
    print("Audio  :", "Exact" if audio_equal else "Different")


COVER_PATH = 'DATASET/Audio/data1_mono.wav'
STEGO_BASE = 'results/STEGOAUDIO'
SEGMENT_LENGTH = 4096
SEGMENT_OVERLAP = 0.5
RANDOM_STATE = 42
TEST_SIZE = 0.3
BALANCE_CLASSES = True


def load_audio_samples(filepath):
    rate, data = scp.read(filepath)
    if data.ndim > 1:
        data = data[:, 0]
    return rate, data.astype(np.float64)


def build_interpolated_reference(cover_samples):
    n = len(cover_samples)
    ref = np.zeros(2 * n - 1)
    ref[::2] = cover_samples
    ref[1::2] = np.floor((cover_samples[:-1] + cover_samples[1:]) / 2.0)
    return ref


def get_segments(samples, seg_len, overlap_ratio=0.5):
    step = int(seg_len * (1 - overlap_ratio))
    if step < 1:
        step = 1
    segments = []
    for start in range(0, len(samples) - seg_len + 1, step):
        segments.append(samples[start:start + seg_len])
    return segments


def extract_statistical_features(segment):
    return [
        np.mean(segment),
        np.std(segment) if np.std(segment) > 0 else 1e-10,
        skew(segment),
        kurtosis(segment),
        np.median(segment),
        np.max(np.abs(segment)),
    ]


def extract_spectral_features(segment):
    fft_mag = np.abs(fft(segment))[:len(segment) // 2]
    if len(fft_mag) < 10:
        return [0.0] * 8

    n = len(fft_mag)
    band_size = n // 4
    bands = [
        fft_mag[:band_size],
        fft_mag[band_size:2 * band_size],
        fft_mag[2 * band_size:3 * band_size],
        fft_mag[3 * band_size:],
    ]
    features = []
    for band in bands:
        if len(band) > 0:
            features.append(np.mean(band))
            features.append(np.std(band) if np.std(band) > 0 else 1e-10)
        else:
            features.extend([0.0, 0.0])
    return features[:8]


def extract_difference_features(segment):
    diff = np.diff(segment.astype(np.int64))
    if len(diff) < 2:
        return [0.0] * 6
    return [
        np.mean(np.abs(diff)),
        np.std(diff) if np.std(diff) > 0 else 1e-10,
        skew(diff),
        np.median(np.abs(diff)),
        np.sum(diff == 0) / len(diff),
        np.sum(np.abs(diff) == 1) / len(diff),
    ]


def extract_histogram_moments(segment, n_bins=32):
    hist, _ = np.histogram(segment, bins=n_bins, density=True)
    hist = hist + 1e-10
    hist = hist / np.sum(hist)
    return [
        np.mean(hist),
        np.std(hist),
        skew(hist),
    ]


def extract_features_for_segment(segment):
    feats = []
    feats.extend(extract_statistical_features(segment))
    feats.extend(extract_spectral_features(segment))
    feats.extend(extract_difference_features(segment))
    feats.extend(extract_histogram_moments(segment))
    return np.array(feats, dtype=np.float64)


def extract_features_for_audio(samples, seg_len, overlap):
    segments = get_segments(samples, seg_len, overlap)
    return np.array([extract_features_for_segment(seg) for seg in segments])


def collect_cover_and_stego_files(cover_path=None):
    cover = cover_path or COVER_PATH
    stego_files = []

    if not os.path.exists(STEGO_BASE):
        return cover, stego_files

    for name in os.listdir(STEGO_BASE):
        if name.startswith('stego_audio'):
            folder = os.path.join(STEGO_BASE, name)
            if os.path.isdir(folder):
                for filename in sorted(os.listdir(folder)):
                    if filename.endswith('.wav'):
                        stego_files.append(os.path.join(folder, filename))

    return cover, stego_files


def build_dataset(cover_path=None):
    from sklearn.utils import resample

    cover_path, stego_paths = collect_cover_and_stego_files(cover_path)

    if not os.path.exists(cover_path):
        raise FileNotFoundError(f"audio cover not found: {cover_path}")

    if not stego_paths:
        raise FileNotFoundError(
            "No stego files found. Please ensure folders like "
            "results/STEGOAUDIO/stego_audio1_payload1/ contain stegoaudio*.wav"
        )

    rate_c, cover_raw = load_audio_samples(cover_path)
    cover_samples = build_interpolated_reference(cover_raw)

    X_list = []
    y_list = []

    X_cover = extract_features_for_audio(cover_samples, SEGMENT_LENGTH, SEGMENT_OVERLAP)
    X_list.append(X_cover)
    y_list.append(np.zeros(len(X_cover), dtype=int))

    for stego_path in stego_paths:
        rate_s, stego_samples = load_audio_samples(stego_path)
        min_len = min(len(cover_samples), len(stego_samples))
        if min_len < SEGMENT_LENGTH:
            continue

        stego_trunc = stego_samples[:min_len]
        X_stego = extract_features_for_audio(stego_trunc, SEGMENT_LENGTH, SEGMENT_OVERLAP)
        X_list.append(X_stego)
        y_list.append(np.ones(len(X_stego), dtype=int))

    X = np.vstack(X_list)
    y = np.concatenate(y_list)

    if BALANCE_CLASSES:
        idx_cover = np.where(y == 0)[0]
        idx_stego = np.where(y == 1)[0]
        n_cover = len(idx_cover)
        n_stego = len(idx_stego)
        if n_stego > n_cover:
            idx_stego_down = resample(
                idx_stego, replace=False, n_samples=n_cover, random_state=RANDOM_STATE
            )
            idx_keep = np.concatenate([idx_cover, idx_stego_down])
            np.random.RandomState(RANDOM_STATE).shuffle(idx_keep)
            X = X[idx_keep]
            y = y[idx_keep]

    return X, y, cover_path, stego_paths


def build_single_pair_dataset(cover_path, stego_path):
    if not os.path.isfile(cover_path):
        raise FileNotFoundError(f"Cover audio not found: {cover_path}")
    if not os.path.isfile(stego_path):
        raise FileNotFoundError(f"Stego audio not found: {stego_path}")

    _, cover_raw = load_audio_samples(cover_path)
    _, stego = load_audio_samples(stego_path)

    cover = (build_interpolated_reference(cover_raw)
             if len(stego) >= len(cover_raw) * 1.8 else cover_raw)
    min_len = min(len(cover), len(stego))
    if min_len < SEGMENT_LENGTH:
        raise ValueError(f"Audio too short for detector (minimum {SEGMENT_LENGTH} samples).")

    X_cover = extract_features_for_audio(cover[:min_len], SEGMENT_LENGTH, SEGMENT_OVERLAP)
    X_stego = extract_features_for_audio(stego[:min_len], SEGMENT_LENGTH, SEGMENT_OVERLAP)
    pair_count = min(len(X_cover), len(X_stego))
    if pair_count < 3:
        raise ValueError("Not enough audio segments for training/testing the detector.")

    X = np.vstack((X_cover[:pair_count], X_stego[:pair_count]))
    X = np.nan_to_num(X, nan=0.0, posinf=0.0, neginf=0.0)
    y = np.concatenate((
        np.zeros(pair_count, dtype=int),
        np.ones(pair_count, dtype=int),
    ))
    return X, y


def build_cover_path(audio_no):
    return f'DATASET/Audio/data{audio_no}_mono.wav'


def evaluate_detector_audio_pair(cover_path, stego_path):
    from sklearn.metrics import accuracy_score, confusion_matrix, f1_score, precision_score, recall_score, roc_auc_score
    from sklearn.model_selection import StratifiedKFold, cross_val_score, train_test_split
    from sklearn.preprocessing import StandardScaler
    from sklearn.svm import SVC

    X, y = build_single_pair_dataset(cover_path, stego_path)
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y
    )

    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    clf = SVC(kernel='rbf', C=1.0, gamma='scale', random_state=RANDOM_STATE)
    clf.fit(X_train_scaled, y_train)
    y_pred = clf.predict(X_test_scaled)

    acc = accuracy_score(y_test, y_pred)
    prec = precision_score(y_test, y_pred, zero_division=0)
    rec = recall_score(y_test, y_pred, zero_division=0)
    f1 = f1_score(y_test, y_pred, zero_division=0)
    try:
        auc = roc_auc_score(y_test, clf.decision_function(X_test_scaled))
    except ValueError:
        auc = 0.0

    class_count = int(np.min(np.bincount(y)))
    folds = min(5, class_count)
    cv_mean = None
    cv_std = None
    if folds >= 2:
        full_scaler = StandardScaler()
        X_scaled = full_scaler.fit_transform(X)
        cv = StratifiedKFold(folds, shuffle=True, random_state=RANDOM_STATE)
        scores = cross_val_score(clf, X_scaled, y, cv=cv)
        cv_mean = float(np.mean(scores))
        cv_std = float(np.std(scores))

    return {
        'accuracy': float(acc),
        'precision': float(prec),
        'recall': float(rec),
        'f1': float(f1),
        'auc': float(auc),
        'cv_mean': cv_mean,
        'cv_std': cv_std,
        'segments_per_class': class_count,
        'confusion_matrix': confusion_matrix(y_test, y_pred),
    }


def run_detector_experiment(cover_path=None):
    from sklearn.metrics import accuracy_score, confusion_matrix, f1_score, precision_score, recall_score, roc_auc_score
    from sklearn.model_selection import StratifiedKFold, cross_val_score, train_test_split
    from sklearn.preprocessing import StandardScaler
    from sklearn.svm import SVC

    print("=" * 70)
    print("   STEGANALYSIS DETECTOR-BASED EVALUATION (ML Benchmark)")
    print(f"   Audio: {cover_path or COVER_PATH}")
    print("=" * 70)

    print("\n[1] Building dataset...")
    X, y, cover_path, stego_paths = build_dataset(cover_path)
    n_cover = np.sum(y == 0)
    n_stego = np.sum(y == 1)
    print(f"    Cover segments : {n_cover}")
    print(f"    Stego segments : {n_stego}")
    print(f"    Total samples  : {len(y)}")
    print(f"    Feature dim    : {X.shape[1]}")
    print(f"    Stego files    : {len(stego_paths)}")

    if n_cover == 0 or n_stego == 0:
        print("\n[ERROR] Not enough data to train/evaluate the detector. Please ensure both cover and stego segments are present.")
        return

    print("\n[2] Split train/test (stratified)...")
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y
    )
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)

    print("\n[3] Training SVM (RBF kernel)...")
    clf = SVC(kernel='rbf', C=1.0, gamma='scale', random_state=RANDOM_STATE)
    clf.fit(X_train_scaled, y_train)
    y_pred = clf.predict(X_test_scaled)

    print("\n[4] Evaluation...")
    acc = accuracy_score(y_test, y_pred)
    prec = precision_score(y_test, y_pred, zero_division=0)
    rec = recall_score(y_test, y_pred, zero_division=0)
    f1 = f1_score(y_test, y_pred, zero_division=0)

    try:
        y_prob = clf.decision_function(X_test_scaled)
        auc = roc_auc_score(y_test, y_prob)
    except Exception:
        auc = 0.0

    cm = confusion_matrix(y_test, y_pred)

    print("\n[5] Cross-validation (5-fold stratified)...")
    X_scaled = scaler.fit_transform(X)
    cv_scores = cross_val_score(
        clf,
        X_scaled,
        y,
        cv=StratifiedKFold(5, shuffle=True, random_state=RANDOM_STATE)
    )
    cv_mean = np.mean(cv_scores)
    cv_std = np.std(cv_scores)

    print("\n" + "=" * 70)
    print("   RESULTS OF DETECTOR-BASED STEGANALYSIS EXPERIMENT")
    print("=" * 70)
    print(f"\n  Confusion Matrix (Test Set):")
    print(f"                   Predicted")
    print(f"                 Cover  Stego")
    print(f"    Actual Cover   {cm[0,0]:>4}    {cm[0,1]:>4}")
    print(f"    Actual Stego   {cm[1,0]:>4}    {cm[1,1]:>4}")
    print(f"\n  Metrics (Test Set):")
    print(f"    Accuracy       : {acc * 100:.2f}%")
    print(f"    Precision      : {prec * 100:.2f}%")
    print(f"    Recall         : {rec * 100:.2f}%")
    print(f"    F1-Score       : {f1 * 100:.2f}%")
    print(f"    AUC-ROC        : {auc * 100:.2f}%")
    print(f"\n  Cross-Validation (5-fold):")
    print(f"    Mean Accuracy  : {cv_mean * 100:.2f}% (+/- {cv_std * 100:.2f}%)")
    print("\n" + "-" * 70)

    return {
        'accuracy': acc,
        'precision': prec,
        'recall': rec,
        'f1': f1,
        'auc': auc,
        'cv_mean': cv_mean,
        'cv_std': cv_std,
        'confusion_matrix': cm,
    }


def calculate_entropy(samples):
    samples_int = samples.astype(np.int64)
    counts = Counter(samples_int)
    total = len(samples_int)
    entropy = 0.0
    for count in counts.values():
        p = count / total
        if p > 0:
            entropy -= p * math.log2(p)
    return entropy


def entropy_analysis(cover_samples, stego_samples):
    if len(stego_samples) >= len(cover_samples) * 1.8:
        ref = build_interpolated_reference(cover_samples)
    else:
        ref = cover_samples
    e_cover = calculate_entropy(ref)
    e_stego = calculate_entropy(stego_samples)
    delta = abs(e_stego - e_cover)
    relative_change = (delta / e_cover * 100) if e_cover > 0 else 0
    return {
        'entropy_cover': e_cover,
        'entropy_stego': e_stego,
        'entropy_diff': delta,
        'entropy_pct': relative_change
    }


def normalized_correlation(cover_samples, stego_samples):
    if len(stego_samples) >= len(cover_samples) * 1.8:
        c = build_interpolated_reference(cover_samples)
    else:
        c = cover_samples

    min_len = min(len(c), len(stego_samples))
    c = c[:min_len]
    s = stego_samples[:min_len]
    numerator = np.sum(c * s)
    denominator = np.sqrt(np.sum(c ** 2) * np.sum(s ** 2))
    return numerator / denominator if denominator > 0 else 0


def evaluate_nc_entropy_audio_pair(cover_path, stego_path):
    if not os.path.isfile(cover_path):
        raise FileNotFoundError(f"Cover audio not found: {cover_path}")
    if not os.path.isfile(stego_path):
        raise FileNotFoundError(f"Stego audio not found: {stego_path}")

    _, cover = load_audio_samples(cover_path)
    _, stego = load_audio_samples(stego_path)

    entropy = entropy_analysis(cover, stego)
    return {
        **entropy,
        'nc': normalized_correlation(cover, stego),
    }


def _detect_total_shares(audio_no, payload_no):
    idx = 0
    while os.path.exists(f'results/STEGOAUDIO/stego_audio{audio_no}_payload{payload_no}/stegoaudio{idx}.wav'):
        idx += 1
    return idx


def _style_header(ws, row, max_col):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border


def _dc(cell, fmt=None):
    from openpyxl.styles import Alignment, Border, Side

    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt


def export_to_excel(all_rows, all_aggregate, excel_path):
    import openpyxl as xl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = xl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Share Details")
    headers = [
        "Audio", "Payload", "Share",
        "Entropy Cover (bits)", "Entropy Stego (bits)",
        "Entropy Diff (bits)", "Entropy Change (%)",
        "NC"
    ]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    _style_header(ws1, 1, len(headers))

    for r, rd in enumerate(all_rows, 2):
        m = rd['m']
        vals = [
            rd['audio'], rd['payload'], rd['share'],
            m['entropy_cover'], m['entropy_stego'],
            m['entropy_diff'], m['entropy_pct'],
            m['nc']
        ]
        fmts = [
            None, None, None,
            '0.000000', '0.000000',
            '0.000000', '0.0000',
            '0.0000000000'
        ]
        for c, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws1.cell(row=r, column=c, value=v)
            _dc(cell, f)

    for c in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 20

    summary_cols = [
        ("Entropy Change (%)", 'entropy_pct', '0.0000', 'min'),
        ("NC", 'nc', '0.0000000000', 'max'),
    ]

    for label, agg_type in [("Summary (Avg)", 'avg'), ("Summary (Best)", 'best'), ("Summary (Worst)", 'worst')]:
        ws = wb.create_sheet(label)
        sh = ["Audio", "Payload", "#Share"] + [s[0] for s in summary_cols]
        for c, h in enumerate(sh, 1):
            ws.cell(row=1, column=c, value=h)
        _style_header(ws, 1, len(sh))

        for r, agg in enumerate(all_aggregate, 2):
            ws.cell(row=r, column=1, value=agg['audio'])
            _dc(ws.cell(row=r, column=1))
            ws.cell(row=r, column=2, value=agg['payload'])
            _dc(ws.cell(row=r, column=2))
            ws.cell(row=r, column=3, value=agg['n_shares'])
            _dc(ws.cell(row=r, column=3))

            for ci, (_, key, fmt, direction) in enumerate(summary_cols):
                vals = [m[key] for m in agg['metrics']]
                if agg_type == 'avg':
                    val = np.mean(vals)
                elif agg_type == 'best':
                    val = max(vals) if direction == 'max' else min(vals)
                else:
                    val = min(vals) if direction == 'max' else max(vals)
                cell = ws.cell(row=r, column=4 + ci, value=val)
                _dc(cell, fmt)

        for c in range(1, len(sh) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 20

    all_shares = sorted(set(r['share'] for r in all_rows))
    all_audios = sorted(set(r['audio'] for r in all_rows))
    all_payloads = sorted(set(r['payload'] for r in all_rows))
    lookup = {(r['audio'], r['payload'], r['share']): r['m'] for r in all_rows}

    pivot_defs = [
        ("Entropy (%)", 'entropy_pct', '0.0000', 1, 'min'),
        ("NC", 'nc', '0.0000000000', 0.999, 'max'),
    ]

    for metric_label, key, fmt, threshold, direction in pivot_defs:
        for share in all_shares:
            sheet_name = f"{metric_label} S{share}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            ws = wb.create_sheet(sheet_name)

            ws.cell(row=1, column=1, value=f"{metric_label} (Share {share})")
            ws.cell(row=1, column=1).font = Font(bold=True, size=12)

            ws.cell(row=2, column=1, value="Audio \\ Payload")
            for ci, payload in enumerate(all_payloads):
                ws.cell(row=2, column=ci + 2, value=f"Payload {payload}")
            _style_header(ws, 2, len(all_payloads) + 1)

            for ri, audio in enumerate(all_audios):
                label_cell = ws.cell(row=ri + 3, column=1, value=f"Audio {audio}")
                label_cell.font = Font(bold=True)
                _dc(label_cell)

                for ci, payload in enumerate(all_payloads):
                    m = lookup.get((audio, payload, share))
                    if m:
                        val = m[key]
                        cell = ws.cell(row=ri + 3, column=ci + 2, value=val)
                        _dc(cell, fmt)
                        if direction == 'min':
                            cell.fill = pass_fill if val < threshold else fail_fill
                        else:
                            cell.fill = pass_fill if val >= threshold else fail_fill
                    else:
                        cell = ws.cell(row=ri + 3, column=ci + 2, value="-")
                        _dc(cell)

            ws.column_dimensions['A'].width = 14
            for ci in range(len(all_payloads)):
                ws.column_dimensions[get_column_letter(ci + 2)].width = 14

    ws_k = wb.create_sheet("Metric Notes")
    info = [
        ["Metric", "Preferred Direction", "PASS Threshold", "Ideal Value", "Formula / Notes"],
        ["Entropy Change (%)", "Lower is better", "< 1%", "0%",
         "Relative difference of Shannon entropy between cover and stego. H(X) = -sum(p(x)*log2(p(x)))"],
        ["NC (Normalized Correlation)", "Higher is better", ">= 0.999", "1.0",
         "Signal correlation. NC = sum(C*S) / sqrt(sum(C^2)*sum(S^2)). Range: -1 to 1"],
    ]
    for ri, row in enumerate(info):
        for ci, val in enumerate(row):
            cell = ws_k.cell(row=ri + 1, column=ci + 1, value=val)
            _dc(cell)
    _style_header(ws_k, 1, 5)
    for c in range(1, 6):
        ws_k.column_dimensions[get_column_letter(c)].width = 32

    os.makedirs(os.path.dirname(excel_path) if os.path.dirname(excel_path) else '.', exist_ok=True)
    wb.save(excel_path)
    print(f"\n  [OK] Results saved to: {excel_path}")


def run_single(audio_no='1', payload_no='1', stego_share=0, cover_path=None):
    cover_path = cover_path or f'DATASET/Audio/data{audio_no}_mono.wav'
    stego_path = f'results/STEGOAUDIO/stego_audio{audio_no}_payload{payload_no}/stegoaudio{stego_share}.wav'

    print("=" * 70)
    print("   STEGANALYSIS (NC and Entropy)")
    print(f"   Cover : {cover_path}")
    print(f"   Stego : {stego_path}")
    print("=" * 70)

    if not os.path.exists(cover_path):
        print(f"[ERROR] Cover audio not found: {cover_path}")
        return
    if not os.path.exists(stego_path):
        print(f"[ERROR] Stego audio not found: {stego_path}")
        return

    _, cover = load_audio_samples(cover_path)
    _, stego = load_audio_samples(stego_path)

    print("\n" + "=" * 70)
    print("  1. ENTROPY ANALYSIS")
    ent = entropy_analysis(cover, stego)
    print(f"  Entropy cover  : {ent['entropy_cover']:.6f} bits")
    print(f"  Entropy stego  : {ent['entropy_stego']:.6f} bits")
    print(f"  Difference     : {ent['entropy_diff']:.6f} bits")
    print(f"  Change         : {ent['entropy_pct']:.4f}%")

    print("\n" + "=" * 70)
    print("  2. NORMALIZED CORRELATION (NC)")
    nc = normalized_correlation(cover, stego)
    print(f"  NC             : {nc:.10f}")

    print("=" * 70)


def run_batch(total_audio=15, total_payload=11, total_shares=None, excel_output=None):
    all_rows = []
    all_aggregate = []

    print("=" * 100)
    print("   BATCH STEGANALYSIS (NC and Entropy) - all shares")
    print("=" * 100)

    header = (f"{'Audio':>6} | {'Payload':>7} | {'Share':>5} | "
              f"{'Entropy%':>9} | {'NC':>14}")
    print(header)
    print("-" * 60)

    for audio in range(1, total_audio + 1):
        for payload in range(1, total_payload + 1):
            cover_path = f'DATASET/Audio/data{audio}_mono.wav'
            if not os.path.exists(cover_path):
                continue

            n_shares = total_shares if total_shares else _detect_total_shares(audio, payload)
            if n_shares == 0:
                continue

            _, cover = load_audio_samples(cover_path)
            combo_metrics = []

            for share in range(n_shares):
                stego_path = f'results/STEGOAUDIO/stego_audio{audio}_payload{payload}/stegoaudio{share}.wav'
                if not os.path.exists(stego_path):
                    continue

                _, stego = load_audio_samples(stego_path)
                ent = entropy_analysis(cover, stego)
                nc = normalized_correlation(cover, stego)

                m = {
                    'entropy_cover': ent['entropy_cover'],
                    'entropy_stego': ent['entropy_stego'],
                    'entropy_diff': ent['entropy_diff'],
                    'entropy_pct': ent['entropy_pct'],
                    'nc': nc,
                }
                combo_metrics.append(m)
                all_rows.append({'audio': audio, 'payload': payload, 'share': share, 'm': m})

                row = (f"{audio:>6} | {payload:>7} | {share:>5} | "
                       f"{m['entropy_pct']:>8.4f}% | "
                       f"{nc:>14.10f}")
                print(row)

            if combo_metrics:
                all_aggregate.append({
                    'audio': audio,
                    'payload': payload,
                    'n_shares': len(combo_metrics),
                    'metrics': combo_metrics
                })

    print("=" * 100)

    if not all_aggregate:
        print("\n[INFO] No stego audio data was found.")
        return

    print("\n" + "=" * 100)
    print("   SUMMARY PER COMBINATION (avg / best / worst)")
    print("   Entropy%: best=min, worst=max | NC: best=max, worst=min")
    print("=" * 100)

    header2 = (f"{'Audio':>6} | {'Payload':>7} | {'#Sh':>4} | "
               f"{'Ent% avg':>9} | {'Ent% best':>9} | {'Ent% worst':>10} | "
               f"{'NC avg':>14} | {'NC best':>14} | {'NC worst':>14}")
    print(header2)
    print("-" * 100)

    for agg in all_aggregate:
        ms = agg['metrics']
        ent_vals = [m['entropy_pct'] for m in ms]
        nc_vals = [m['nc'] for m in ms]
        row2 = (f"{agg['audio']:>6} | {agg['payload']:>7} | {agg['n_shares']:>4} | "
                f"{np.mean(ent_vals):>8.4f}% | {min(ent_vals):>8.4f}% | {max(ent_vals):>9.4f}% | "
                f"{np.mean(nc_vals):>14.10f} | {max(nc_vals):>14.10f} | {min(nc_vals):>14.10f}")
        print(row2)

    print("=" * 100)

    if excel_output is None:
        n_str = str(total_shares) if total_shares else 'auto'
        excel_output = f'results/STEGANALYSIS/steganalysis_{n_str}_shares.xlsx'

    export_to_excel(all_rows, all_aggregate, excel_output)
    return all_rows, all_aggregate


def run_detector_ml():
    print("=" * 70)
    print("  DETECTOR-BASED STEGANALYSIS (ML Benchmark)")
    print("=" * 70)

    print("Select stego_audio[X]_payload[Y] to analyze:")
    audio_no = input("Enter audio number [X]: ").strip() or "1"
    payload_no = input("Enter payload number [Y]: ").strip() or "1"
    if not audio_no.isdigit() or not payload_no.isdigit():
        raise ValueError("Audio number and payload number must be integers.")

    cover_path = build_cover_path(audio_no)
    run_detector_experiment(cover_path)


def run_nc_entropy_single():
    print("=" * 70)
    print("  NC AND ENTROPY ANALYSIS (single pair)")
    print("=" * 70)

    print("Select stego_audio[X]_payload[Y] to analyze:")
    audio_no = input("Enter audio number [X]: ").strip() or "1"
    payload_no = input("Enter payload number [Y]: ").strip() or "1"
    if not audio_no.isdigit() or not payload_no.isdigit():
        raise ValueError("Audio number and payload number must be integers.")

    cover_path = build_cover_path(audio_no)
    share_input = input("Enter stegoaudio number (default=0): ").strip()
    share_no = int(share_input) if share_input else 0
    run_single(audio_no, payload_no, share_no, cover_path)


def run_nc_entropy_batch():
    print("=" * 70)
    print("  NC AND ENTROPY ANALYSIS (batch + Excel)")
    print("=" * 70)

    audio_input = input("Total audio files (default=15): ").strip()
    payload_input = input("Total payloads (default=11): ").strip()
    shares_input = input("Total shares / n (blank = auto-detect): ").strip()
    run_batch(
        total_audio=int(audio_input) if audio_input else 15,
        total_payload=int(payload_input) if payload_input else 11,
        total_shares=int(shares_input) if shares_input else None,
    )


def validate_share_values(total_shares, min_shares):
    if total_shares < 1 or not 1 <= min_shares <= total_shares:
        raise ValueError("Value of shares must satisfy 1 <= k <= n.")


def run_cli_embedding(args):
    audio_file = 'DATASET/Audio/data{}_mono.wav'.format(args.audio)
    payload_file = 'DATASET/Payload/payload{}.txt'.format(args.payload)
    output_base = args.output_base or 'results/STEGOAUDIO/stego_audio{}_payload{}/stegoaudio'.format(
        args.audio,
        args.payload
    )

    if not os.path.isfile(audio_file):
        raise FileNotFoundError("Audio file not found: {}".format(audio_file))
    if not os.path.isfile(payload_file):
        raise FileNotFoundError("Payload file not found: {}".format(payload_file))
    validate_share_values(args.shares, args.min_shares)

    result = run_single_embedding(
        payload_file,
        audio_file,
        args.shares,
        args.min_shares,
        output_base
    )

    print("\nEmbedding completed successfully.")
    print("Output folder:", result['output_dir'])
    print("Peak memory:", result['peak_memory_mb'], "MB")
    print("Embedding runtime:", result['runtime'])


def run_cli_extraction(args):
    validate_share_values(args.shares, args.min_shares)

    stego_audio_base = 'results/STEGOAUDIO/stego_audio{}_payload{}/stegoaudio'.format(
        args.audio,
        args.payload
    )
    output_dir = args.output_dir or 'results/EXTRACTED/stego_audio{}_payload{}'.format(
        args.audio,
        args.payload
    )

    missing_files = [
        '{}{}.wav'.format(stego_audio_base, index)
        for index in range(args.shares)
        if not os.path.isfile('{}{}.wav'.format(stego_audio_base, index))
    ]
    if missing_files:
        raise FileNotFoundError(
            "Incomplete stego-audio file. First file not found: {}".format(
                missing_files[0]
            )
        )

    if args.share_indices:
        selected_shares = args.share_indices
        if len(selected_shares) != args.min_shares:
            raise ValueError("--share-indices must contain exactly k values.")
        if any(index < 0 or index >= args.shares for index in selected_shares):
            raise ValueError("--share-indices values must be between 0 and n-1.")
    else:
        selected_shares = random.sample(range(args.shares), args.min_shares)

    frame_rate, stego_sample = extraction_sampling(stego_audio_base, selected_shares)
    payload_output, cover_audio_output = extract_payload_and_audio(stego_sample, output_dir)

    print("\nExtraction completed successfully.")
    print("Share used:", selected_shares)
    print("Extracted payload:", payload_output)
    print("Extracted audio:", cover_audio_output)


def run_cli_compare(args):
    original_payload = 'DATASET/Payload/payload{}.txt'.format(args.payload)
    extracted_payload = 'results/EXTRACTED/stego_audio{}_payload{}/payload.txt'.format(
        args.audio,
        args.payload
    )
    original_audio = 'DATASET/Audio/data{}_mono.wav'.format(args.audio)
    extracted_audio = 'results/EXTRACTED/stego_audio{}_payload{}/audio.wav'.format(
        args.audio,
        args.payload
    )

    for filepath in (
        original_payload, extracted_payload, original_audio, extracted_audio
    ):
        require_file(filepath)

    payload_equal = compare_data(
        read_payload(original_payload),
        read_payload(extracted_payload),
        'payload'
    )
    audio_equal = compare_data(
        sampling_audio_for_compare(original_audio),
        sampling_audio_for_compare(extracted_audio),
        'audio'
    )

    print("\nComparison results:")
    print("Payload:", "Exact" if payload_equal else "Different")
    print("Audio  :", "Exact" if audio_equal else "Different")


def run_cli_quality(args):
    file_audio = 'DATASET/Audio/data{}_mono.wav'.format(args.audio)
    clone_filename = args.clone_file or 'results/CLONING/data_clone_audio{}.wav'.format(args.audio)
    file_stego_audio = 'results/STEGOAUDIO/stego_audio{}_payload{}/stegoaudio{}.wav'.format(
        args.audio,
        args.payload,
        args.share
    )

    if not os.path.isfile(file_audio):
        raise FileNotFoundError("Cover audio not found: {}".format(file_audio))
    if not os.path.isfile(file_stego_audio):
        raise FileNotFoundError("Stego audio not found: {}".format(file_stego_audio))

    result = run_single_quality(file_audio, file_stego_audio, clone_filename)

    print("\nQuality evaluation completed successfully.")
    print("mse  :", result['mse'])
    print("snr  :", result['snr'])
    print("psnr :", result['psnr'])


def run_cli_detector(args):
    cover_path = args.cover_path or build_cover_path(args.audio)
    run_detector_experiment(cover_path)


def run_cli_nc_single(args):
    cover_path = args.cover_path or build_cover_path(args.audio)
    run_single(args.audio, args.payload, args.share, cover_path)


def run_cli_nc_batch(args):
    run_batch(
        total_audio=args.total_audio,
        total_payload=args.total_payload,
        total_shares=args.shares,
        excel_output=args.excel_output,
    )


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="StegoShare reversible audio steganography toolkit."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    embedding_parser = subparsers.add_parser("embedding", help="Embed a payload into cover audio.")
    embedding_parser.add_argument("--audio", required=True, help="Audio number X.")
    embedding_parser.add_argument("--payload", required=True, help="Payload number Y.")
    embedding_parser.add_argument("--shares", required=True, type=int, help="Total shares n.")
    embedding_parser.add_argument("--min-shares", required=True, type=int, help="Minimum shares k.")
    embedding_parser.add_argument("--output-base", help="Optional output path without share index and .wav extension.")
    embedding_parser.set_defaults(func=run_cli_embedding)

    extraction_parser = subparsers.add_parser("extracting", help="Extract payload and audio from stego shares.")
    extraction_parser.add_argument("--audio", required=True, help="Audio number X.")
    extraction_parser.add_argument("--payload", required=True, help="Payload number Y.")
    extraction_parser.add_argument("--shares", required=True, type=int, help="Total shares n.")
    extraction_parser.add_argument("--min-shares", required=True, type=int, help="Minimum shares k.")
    extraction_parser.add_argument("--share-indices", nargs="+", type=int, help="Optional exact share indices to use.")
    extraction_parser.add_argument("--output-dir", help="Optional extraction output directory.")
    extraction_parser.set_defaults(func=run_cli_extraction)

    compare_parser = subparsers.add_parser("compare", help="Compare original and extracted payload/audio.")
    compare_parser.add_argument("--audio", required=True, help="Audio number X.")
    compare_parser.add_argument("--payload", required=True, help="Payload number Y.")
    compare_parser.set_defaults(func=run_cli_compare)

    quality_parser = subparsers.add_parser("quality", help="Evaluate stego-audio quality.")
    quality_parser.add_argument("--audio", required=True, help="Audio number X.")
    quality_parser.add_argument("--payload", required=True, help="Payload number Y.")
    quality_parser.add_argument("--share", default=0, type=int, help="Stego-audio share index.")
    quality_parser.add_argument("--clone-file", help="Optional cloned cover audio output path.")
    quality_parser.set_defaults(func=run_cli_quality)

    detector_parser = subparsers.add_parser("detector", help="Run SVM detector-based steganalysis.")
    detector_parser.add_argument("--audio", default="1", help="Audio number X.")
    detector_parser.add_argument("--cover-path", help="Optional explicit cover audio path.")
    detector_parser.set_defaults(func=run_cli_detector)

    nc_single_parser = subparsers.add_parser("nc-single", help="Run single NC and entropy analysis.")
    nc_single_parser.add_argument("--audio", default="1", help="Audio number X.")
    nc_single_parser.add_argument("--payload", default="1", help="Payload number Y.")
    nc_single_parser.add_argument("--share", default=0, type=int, help="Stego-audio share index.")
    nc_single_parser.add_argument("--cover-path", help="Optional explicit cover audio path.")
    nc_single_parser.set_defaults(func=run_cli_nc_single)

    nc_batch_parser = subparsers.add_parser("nc-batch", help="Run batch NC and entropy analysis.")
    nc_batch_parser.add_argument("--total-audio", default=15, type=int, help="Total cover audio files.")
    nc_batch_parser.add_argument("--total-payload", default=11, type=int, help="Total payload files.")
    nc_batch_parser.add_argument("--shares", type=int, help="Total shares n. Leave blank for auto-detection.")
    nc_batch_parser.add_argument("--excel-output", help="Optional Excel output path.")
    nc_batch_parser.set_defaults(func=run_cli_nc_batch)

    return parser


def run_cli(argv=None):
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    args.func(args)


def run_evaluation():
    show_menu_evaluation()
    choice = input("Choose (1-5): ").strip()

    menu = EVALUATION_MENU.get(choice)
    if menu is None:
        print("Invalid choice. Please enter a number between 1 and 5.")
        return

    name, target = menu
    print("\nRunning {}...".format(name))

    if target == "detector_ml":
        run_detector_ml()
    elif target == "single_compare":
        run_single_compare()
    elif target == "quality_evaluation":
        run_quality_evaluation()
    elif target == "nc_entropy_single":
        run_nc_entropy_single()
    elif target == "nc_entropy_batch":
        run_nc_entropy_batch()
    else:
        print("Invalid evaluation target.")


def main():
    while True:
        show_menu()
        choice = input("Choose (1-4): ").strip()

        if choice == "4":
            print("Program Closed.")
            break

        menu = MENU.get(choice)
        if menu is None:
            print("Invalid choice. Please enter a number between 1 and 4.")
            continue

        name, name_modul = menu
        print("\nRunning {}...".format(name))

        try:
            if choice == "3":
                run_evaluation()
            elif choice == "1":
                run_embedding()
            elif choice == "2":
                run_extraction()
            else:
                print("Invalid menu target.")
        except KeyboardInterrupt:
            print("\nCanceled by user.")
        except Exception as error:
            print("\nFailed to run {}: {}".format(name, error))


if __name__ == "__main__":
    if len(sys.argv) > 1:
        run_cli()
    else:
        main()
